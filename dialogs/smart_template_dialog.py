# dialogs/smart_template_dialog.py
"""
Smart Template Dialog v2.0 - Advanced Import System
===================================================
Features:
- Anchor-based templates (resistant to row/col shifts)
- Multi-strategy field detection (keyword, fuzzy, context, proximity)
- Full sub-code name resolution
- Adaptive learning from user corrections
- Context-aware scoring
- Robust value extraction with radius search
- Smart sheet routing with content analysis
"""

import os
import re
import json
import math
import logging
from datetime import datetime, date as dt_date, time as dt_time
from typing import Dict, List, Any, Optional, Tuple, Set
from difflib import SequenceMatcher
from collections import defaultdict

from PySide6.QtWidgets import *
from PySide6.QtCore import *
from PySide6.QtGui import *
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

MAX_PREVIEW_ROWS = 200
MAX_PREVIEW_COLS = 80
MAX_SCAN_ROWS = 500
MAX_SCAN_COLS = 150

# =====================================================
# Comprehensive Code Maps
# =====================================================
MAIN_CODE_MAP = {
    "1": "Rig Up/ Tear Down / Move",
    "2": "Drilling",
    "3": "Reaming",
    "4": "Coring",
    "5": "Circulate & Condition",
    "6": "Trips",
    "7": "Service/ Maintain Rig",
    "8": "Repair Rig",
    "9": "Replacing Drill Line",
    "10": "Deviation Survey",
    "11": "Logging",
    "12": "Run Casing/ Liner",
    "13": "Cementing",
    "14": "Wait on Cement",
    "15": "Rig Up/Down BOP",
    "16": "Drill Stem Test",
    "17": "Fishing",
    "18": "Specialized Directional Work",
    "19": "Operation Status (Waiting)",
    "20": "Safety",
    "21": "Perforating",
    "22": "Completion/XMT",
    "23": "Treating",
    "24": "Swabbing",
    "25": "Surface Testing",
    "26": "Well Control",
    "27": "Other",
}

SUB_CODE_MAP = {
    # 1 - Rig Up/Down/Move
    "1.1": "Rig Moving/Positioning",
    "1.2": "Rig Up",
    "1.3": "Rig Down",
    "1.4": "Rig Up/Down Completion Equipment",
    "1.5": "Rig Up/Down Testing Equipment",
    "1.6": "Skid Rig",
    # 2 - Drilling
    "2.1": "Vertical Drilling",
    "2.2": "Directional Drilling (Rotating)",
    "2.3": "Directional Drilling (Sliding)",
    "2.4": "Drill Out Plug/Cement",
    "2.5": "Milling/Under-reaming While Drilling",
    # 3 - Reaming
    "3.1": "Reaming / Back Reaming",
    "3.2": "Coring Reaming",
    "3.3": "Hole Opening / Under-reaming",
    "3.4": "Drill Out Cement/ Shoe track",
    "3.5": "Wash Down",
    # 4 - Coring
    "4.1": "Conventional Coring",
    "4.2": "Orientation Coring",
    "4.3": "Sponge Coring",
    # 5 - Circulate
    "5.1": "Hole displacement",
    "5.2": "Circulate/ Condition Mud",
    "5.3": "Circulate for Sampling",
    "5.4": "Break Circulation",
    "5.5": "Lost Circulation Treatment",
    # 6 - Trips
    "6.1": "Trip In (Dry)",
    "6.2": "Trip Out (Dry)",
    "6.3": "Trip In (Wet)",
    "6.4": "Trip Out (Wet)",
    "6.5": "Run in Hole",
    "6.6": "Pull Out Of Hole",
    "6.7": "Pick Up/Lay Down BHA",
    "6.8": "Wiper/ Condition Trip",
    "6.9": "Short Trip",
    "6.10": "Handle Tubulars",
    # 7 - Service/Maintain
    "7.1": "Service Equipment",
    "7.2": "Maintain Equipment",
    "7.3": "Service Mud System",
    "7.4": "Slip & Cut Drill Line",
    # 8 - Repair
    "8.1": "Repair Rig Equipment",
    "8.2": "Repair Downhole Equipment",
    "8.3": "Repair Surface Equipment",
    # 9 - Drill Line
    "9.1": "Replace/Slip Drill Line",
    # 10 - Survey
    "10.1": "MWD Survey",
    "10.2": "Gyro Survey",
    "10.3": "Single Shot Survey",
    "10.4": "Multi Shot Survey",
    # 11 - Logging
    "11.1": "Run in Hole Logging Tools",
    "11.2": "Wire line logging",
    "11.3": "Logging While Drilling",
    "11.4": "Pull Out Logging Tools",
    "11.5": "Mud Logging",
    # 12 - Casing/Liner
    "12.1": "CSG Prep/Handling",
    "12.2": "CSG Running",
    "12.3": "Run Liner Hanger",
    "12.4": "CSG Accessories",
    "12.5": "Liner Running",
    "12.6": "Run Tieback",
    # 13 - Cementing
    "13.1": "Casing/ Liner Cementing",
    "13.2": "Plug Cementing",
    "13.3": "Squeeze CMT",
    "13.4": "Wait on Cement (WOC)",
    "13.5": "Cement Evaluation",
    # 14 - WOC
    "14.1": "Wait on Cement",
    # 15 - BOP
    "15.1": "Rig Up BOP",
    "15.2": "Test BOP",
    "15.3": "Rig Down BOP",
    "15.4": "BOP Maintenance",
    # 16 - DST
    "16.1": "Run DST Tools",
    "16.2": "Drill Stem Test",
    "16.3": "Pull DST Tools",
    # 17 - Fishing
    "17.1": "Fishing Job",
    "17.2": "Jar/Work String",
    "17.3": "Work on Stuck",
    "17.4": "Milling (Fish)",
    "17.5": "Sidetrack Preparation",
    # 18 - Directional
    "18.1": "Run Motor/RSS",
    "18.2": "MWD/LWD Operations",
    "18.3": "Orientation",
    "18.4": "Jar Operations",
    # 19 - Waiting
    "19.1": "Waiting on Client",
    "19.2": "Waiting on Equipment",
    "19.3": "Waiting on Material",
    "19.4": "Waiting on Third Party",
    "19.5": "Waiting on Weather",
    "19.6": "Waiting on Orders",
    "19.7": "Waiting on Daylight",
    "19.8": "Waiting on Rig Repair",
    # 20 - Safety
    "20.1": "Pre Job Safety Meeting (PJSM)",
    "20.2": "Safety Drill",
    "20.3": "H2S Drill",
    "20.4": "Fire Drill",
    "20.5": "Man Overboard Drill",
    # 21 - Perforating
    "21.1": "Run Perforating Guns",
    "21.2": "Perforate",
    "21.3": "Pull Perforating Guns",
    # 22 - Completion
    "22.1": "Run Completion String",
    "22.2": "Set Packer",
    "22.3": "Install XMT",
    "22.4": "Gravel Pack",
    "22.5": "Run Screens",
    # 23 - Treating
    "23.1": "Acidize",
    "23.2": "Fracture",
    "23.3": "Chemical Treatment",
    # 24 - Swabbing
    "24.1": "Swab Well",
    # 25 - Surface Testing
    "25.1": "Flow Well to Test",
    "25.2": "Well Testing",
    "25.3": "Shut In Well",
    # 26 - Well Control
    "26.1": "Shut In Well (Kick)",
    "26.2": "Kill Well",
    "26.3": "FIT/ LOT",
    "26.4": "Flow Check",
    "26.5": "Strip/Snub",
    # 27 - Other
    "27.1": "Other Operations",
    "27.2": "Rig Inspection",
    "27.3": "Government Inspection",
}

# Reverse map: name -> code for lookup
MAIN_CODE_REVERSE = {v.lower(): k for k, v in MAIN_CODE_MAP.items()}
SUB_CODE_REVERSE = {v.lower(): k for k, v in SUB_CODE_MAP.items()}

# NPT contractor mapping
NPT_CONTRACTOR_MAP = {
    "rr": "Rig Contractor",
    "rr-pump": "Rig Contractor",
    "rr-draw": "Rig Contractor",
    "rr-rotary": "Rig Contractor",
    "rr-elec": "Rig Contractor",
    "rr-hyd": "Rig Contractor",
    "rr-mast": "Rig Contractor",
    "rr-bop": "Rig Contractor",
    "rr-pipe": "Rig Contractor",
    "f-mwd": "Directional Company",
    "f-motor": "Directional Company",
    "f-rss": "Directional Company",
    "f-lwd": "Directional Company",
    "f-mud": "Mud Company",
    "f-cement": "Cementing Company",
    "f-wireline": "Wireline Company",
    "f-logging": "Logging Company",
    "f-casing": "Casing Company",
    "f-bit": "Bit Supplier",
    "f-screen": "Screen Supplier",
    "f-packer": "Completion Company",
    "f-tool": "Tool Company",
    "w-weather": "Weather",
    "w-client": "Client",
    "w-material": "Logistics",
    "w-fuel": "Logistics",
    "w-water": "Logistics",
    "w-service": "Service Company",
    "w-order": "Client",
    "w-daylight": "Operations",
    "w-rig": "Rig Contractor",
    "t-stuck": "Operations",
    "t-kick": "Operations",
    "t-loss": "Operations",
    "t-hole": "Operations",
    "t-twist": "Operations",
    "t-wellbore": "Operations",
}


# =====================================================
# Field Patterns v2 - Enhanced with synonyms and context
# =====================================================
FIELD_PATTERNS = {
    # Well Info
    "well_info.name": {
        "keywords": ["well name", "well:", "well", "نام چاه", "well designation"],
        "synonyms": ["wellname", "well_name", "well id"],
        "type": "str",
        "context_group": "well_header",
        "priority": 10,
    },
    "well_info.field_name": {
        "keywords": ["field:", "field name", "field", "میدان", "oil field"],
        "synonyms": ["fieldname", "field_name", "area"],
        "type": "str",
        "context_group": "well_header",
        "priority": 9,
    },
    "well_info.well_type": {
        "keywords": ["well type", "نوع چاه", "type of well"],
        "synonyms": ["welltype", "well_type"],
        "type": "str",
        "context_group": "well_header",
    },
    "well_info.well_shape": {
        "keywords": ["well shape", "شکل چاه", "well profile", "well trajectory type"],
        "synonyms": ["wellshape", "profile"],
        "type": "str",
        "context_group": "well_header",
    },
    "well_info.client": {
        "keywords": ["client:", "client", "کارفرما", "company", "oil company"],
        "synonyms": ["operator company", "client name"],
        "type": "str",
        "context_group": "personnel",
        "priority": 8,
    },
    "well_info.client_rep": {
        "keywords": ["client rep", "client representative", "company rep"],
        "synonyms": ["client_rep", "company representative"],
        "type": "str",
        "context_group": "personnel",
    },
    "well_info.operator": {
        "keywords": ["operator", "اپراتور", "operating company"],
        "synonyms": ["operator name", "op company"],
        "type": "str",
        "context_group": "personnel",
    },
    "well_info.rig_name": {
        "keywords": ["rig name", "rig:", "نام دکل", "rig", "drilling rig"],
        "synonyms": ["rigname", "rig_name", "rig no", "rig number"],
        "type": "str",
        "context_group": "well_header",
        "priority": 8,
    },
    "well_info.drilling_contractor": {
        "keywords": ["drilling contractor", "contractor", "پیمانکار حفاری"],
        "synonyms": ["drill contractor", "rig contractor", "drilling company"],
        "type": "str",
        "context_group": "personnel",
    },
    "well_info.gle_msl": {
        "keywords": ["gle-msl", "gle msl", "ground level elevation"],
        "synonyms": ["gle", "ground elevation"],
        "type": "float",
        "context_group": "elevations",
    },
    "well_info.rte_msl": {
        "keywords": ["rte-msl", "rte msl", "rotary table elevation"],
        "synonyms": ["rte", "rt elevation", "kb elevation", "kb-msl"],
        "type": "float",
        "context_group": "elevations",
    },
    "well_info.gle_rte": {
        "keywords": ["gle-rte", "rt - wh", "rt-wh", "rte-gle"],
        "synonyms": ["rt to wh", "rt - gl"],
        "type": "float",
        "context_group": "elevations",
    },
    "well_info.target_depth": {
        "keywords": [
            "estimated final depth", "target depth", "total depth",
            "final depth", "td", "efd",
        ],
        "synonyms": ["target_depth", "planned td", "proposed td", "efd md"],
        "type": "float",
        "valid_range": (50, 15000),
        "context_group": "depths",
        "priority": 9,
    },
    "well_info.water_depth": {
        "keywords": ["water depth", "عمق آب", "sea depth"],
        "synonyms": ["waterdepth", "water_depth"],
        "type": "float",
        "valid_range": (0, 5000),
        "context_group": "depths",
    },
    "well_info.section_name": {
        "keywords": ["hole section", "section:", "section", "hole size"],
        "synonyms": ["hole_section", "current section", "sect"],
        "type": "str",
        "context_group": "well_header",
        "priority": 8,
    },
    "well_info.formation": {
        "keywords": ["formation:", "formation", "current formation", "fm"],
        "synonyms": ["geological formation", "geo formation"],
        "type": "str",
        "context_group": "geology",
    },
    "well_info.operation_manager": {
        "keywords": ["operation manager", "operations manager"],
        "synonyms": ["ops manager", "op manager"],
        "type": "str",
        "context_group": "personnel",
    },
    "well_info.superintendent": {
        "keywords": ["superintendent", "drilling superintendent"],
        "synonyms": ["supt", "drill supt"],
        "type": "str",
        "context_group": "personnel",
    },
    "well_info.supervisor_day": {
        "keywords": ["supervisor (day)", "day supervisor", "supv day"],
        "synonyms": ["supervisor day", "day supv", "d/s"],
        "type": "str",
        "context_group": "personnel",
    },
    "well_info.supervisor_night": {
        "keywords": ["supervisor (night)", "night supervisor", "supv night"],
        "synonyms": ["supervisor night", "night supv", "n/s"],
        "type": "str",
        "context_group": "personnel",
    },
    "well_info.geologist1": {
        "keywords": ["well site geologist", "geologist", "wsg"],
        "synonyms": ["geo", "wellsite geologist", "site geologist"],
        "type": "str",
        "context_group": "personnel",
    },
    "well_info.tool_pusher_day": {
        "keywords": ["tool pusher", "toolpusher", "driller"],
        "synonyms": ["tp", "tool_pusher", "day driller"],
        "type": "str",
        "context_group": "personnel",
    },
    "well_info.lta_day": {
        "keywords": ["lta (day)", "lta day", "lta", "last time accident"],
        "synonyms": ["lta_day", "days since lta"],
        "type": "int",
        "context_group": "safety",
    },
    "well_info.actual_rig_days": {
        "keywords": ["actual rig day", "rig days", "days on well"],
        "synonyms": ["rig_days", "actual days"],
        "type": "float",
        "context_group": "timing",
    },
    "well_info.kop1": {
        "keywords": ["kop", "kop #1", "kick off point", "kop1"],
        "synonyms": ["kick off", "kick_off_point"],
        "type": "float",
        "valid_range": (0, 15000),
        "context_group": "depths",
    },
    "well_info.rig_heading": {
        "keywords": ["rig heading", "heading", "rig azimuth"],
        "synonyms": ["rig_heading", "azimuth"],
        "type": "float",
        "valid_range": (0, 360),
        "context_group": "well_header",
    },
    "well_info.spud_date": {
        "keywords": ["spud", "spud date", "spud :"],
        "synonyms": ["spud_date", "date spudded"],
        "type": "date",
        "context_group": "dates",
        "priority": 8,
    },
    "well_info.report_date": {
        "keywords": ["report date", "report:", "date:", "date"],
        "synonyms": ["report_date", "rpt date"],
        "type": "date",
        "context_group": "dates",
        "priority": 9,
    },
    "well_info.start_hole_date": {
        "keywords": ["hole section start", "section start date", "start date"],
        "synonyms": ["section_start", "hole start"],
        "type": "date",
        "context_group": "dates",
    },
    "well_info.latitude": {
        "keywords": ["latitude", "lat"],
        "synonyms": ["lat.", "lattitude"],
        "type": "float",
        "valid_range": (-90, 90),
        "context_group": "coordinates",
        "min_keyword_length": 3,
    },
    "well_info.longitude": {
        "keywords": ["longitude", "long", "lon"],
        "synonyms": ["long.", "lng"],
        "type": "float",
        "valid_range": (-180, 180),
        "context_group": "coordinates",
        "min_keyword_length": 3,
    },
    "well_info.northing": {
        "keywords": ["northing", "northing (m)"],
        "synonyms": ["north", "utm northing"],
        "type": "float",
        "context_group": "coordinates",
    },
    "well_info.easting": {
        "keywords": ["easting", "easting (m)"],
        "synonyms": ["east", "utm easting"],
        "type": "float",
        "context_group": "coordinates",
    },
    # Daily Report
    "daily_report.report_number": {
        "keywords": ["report no", "report number", "rpt no", "report #"],
        "synonyms": ["rpt number", "ddr no", "report_no"],
        "type": "int",
        "context_group": "report_header",
        "priority": 10,
    },
    "daily_report.report_date": {
        "keywords": ["report date", "date", "rpt date"],
        "synonyms": ["report_date", "ddr date"],
        "type": "date",
        "context_group": "report_header",
        "priority": 10,
    },
    "daily_report.depth_0000": {
        "keywords": [
            "md (m)@ 0:00", "depth @ 00:00", "@ 0:00",
            "depth at 00:00", "md @ 00:00",
        ],
        "synonyms": ["depth_0000", "midnight depth", "0:00 depth"],
        "type": "float",
        "valid_range": (0, 15000),
        "context_group": "report_depths",
        "priority": 9,
    },
    "daily_report.depth_0600": {
        "keywords": [
            "md (m)@ 6:00", "depth @ 06:00", "@ 6:00",
            "depth at 06:00", "md @ 06:00",
        ],
        "synonyms": ["depth_0600", "morning depth", "6:00 depth"],
        "type": "float",
        "valid_range": (0, 15000),
        "context_group": "report_depths",
        "priority": 9,
    },
    "daily_report.depth_2400": {
        "keywords": [
            "md (m)@ 24:00", "depth @ 24:00", "@ 24:00",
            "depth at 24:00", "md @ 24:00",
        ],
        "synonyms": ["depth_2400", "end depth", "24:00 depth"],
        "type": "float",
        "valid_range": (0, 15000),
        "context_group": "report_depths",
        "priority": 9,
    },
    "daily_report.rig_day": {
        "keywords": ["rig day", "day number", "day no", "day #"],
        "synonyms": ["rig_day", "drilling day"],
        "type": "int",
        "context_group": "report_header",
    },
    "daily_report.summary": {
        "keywords": [
            "summary of activities", "summary of operation",
            "remarks", "summary", "operations summary",
        ],
        "synonyms": ["daily summary", "activity summary"],
        "type": "text",
        "context_group": "report_text",
    },
    "daily_report.forecast": {
        "keywords": [
            "operation forecast", "forecast for next",
            "planned operations", "forecast", "next 24 hrs",
        ],
        "synonyms": ["plan", "next operations"],
        "type": "text",
        "context_group": "report_text",
    },
    # Mud Report
    "mud_report.mud_type": {
        "keywords": ["mud type", "نوع گل", "fluid type", "drilling fluid"],
        "synonyms": ["mud_type", "fluid system"],
        "type": "str",
        "context_group": "mud_props",
    },
    "mud_report.mw": {
        "keywords": [
            "mud weight", "mw", "وزن گل", "density",
            "mud wt", "mud density",
        ],
        "synonyms": ["mud_weight", "m.w.", "m.w"],
        "type": "float",
        "valid_range": (30, 200),
        "context_group": "mud_props",
        "priority": 9,
        "min_keyword_length": 2,
    },
    "mud_report.pv": {
        "keywords": ["pv (cp)", "pv", "plastic viscosity"],
        "synonyms": ["p.v.", "plastic visc"],
        "type": "float",
        "valid_range": (0, 150),
        "context_group": "mud_rheology",
        "min_keyword_length": 2,
    },
    "mud_report.yp": {
        "keywords": ["yp", "yield point", "yp (lb/100ft²)"],
        "synonyms": ["y.p.", "yield pt"],
        "type": "float",
        "valid_range": (0, 150),
        "context_group": "mud_rheology",
        "min_keyword_length": 2,
    },
    "mud_report.funnel_vis": {
        "keywords": ["funnel vis", "funnel viscosity", "fv", "marsh funnel"],
        "synonyms": ["funnel_vis", "fv (sec)", "marsh"],
        "type": "float",
        "valid_range": (20, 500),
        "context_group": "mud_rheology",
    },
    "mud_report.gel_10s": {
        "keywords": ["gel 10 s", "gel 10s", "gel 10sec", "10s gel"],
        "synonyms": ["gel_10s", "gel10s", "10\" gel"],
        "type": "float",
        "context_group": "mud_rheology",
    },
    "mud_report.gel_10m": {
        "keywords": ["gel 10 min", "gel 10m", "gel 10min", "10m gel"],
        "synonyms": ["gel_10m", "gel10m", "10' gel"],
        "type": "float",
        "context_group": "mud_rheology",
    },
    "mud_report.fl": {
        "keywords": ["fluid loss", "api fluid loss", "api fl", "filtrate"],
        "synonyms": ["fl", "f.l.", "filtration"],
        "type": "float",
        "valid_range": (0, 100),
        "context_group": "mud_filtration",
    },
    "mud_report.ph": {
        "keywords": ["ph"],
        "synonyms": ["p.h.", "acidity"],
        "type": "float",
        "min_keyword_length": 2,
        "valid_range": (0, 14),
        "context_group": "mud_chemistry",
    },
    "mud_report.temperature": {
        "keywords": ["flow line temp", "temperature", "temp", "flt"],
        "synonyms": ["flowline temp", "flow_line_temp", "mud temp"],
        "type": "float",
        "valid_range": (0, 250),
        "context_group": "mud_props",
    },
    "mud_report.chloride": {
        "keywords": ["chloride", "cl-", "nacl"],
        "synonyms": ["chlorides", "cl"],
        "type": "float",
        "valid_range": (0, 300000),
        "context_group": "mud_chemistry",
    },
    "mud_report.solid_percent": {
        "keywords": ["solid (% vol)", "solid", "solids", "total solids"],
        "synonyms": ["solid_percent", "solids %", "% solids"],
        "type": "float",
        "context_group": "mud_solids",
    },
    "mud_report.oil_percent": {
        "keywords": ["oil / water", "oil/water", "oil water ratio", "o/w"],
        "synonyms": ["oil_percent", "oil %", "oil ratio"],
        "type": "float",
        "context_group": "mud_solids",
    },
    "mud_report.cake_thickness": {
        "keywords": ["cake thickness", "filter cake", "cake"],
        "synonyms": ["cake", "mud cake"],
        "type": "float",
        "context_group": "mud_filtration",
    },
    "mud_report.volume_hole": {
        "keywords": ["vol. in hole", "vol in hole", "volume in hole"],
        "synonyms": ["volume_hole", "hole volume"],
        "type": "float",
        "context_group": "mud_volume",
    },
    "mud_report.loss_surface": {
        "keywords": ["lost at surface", "surface loss", "loss surface"],
        "synonyms": ["loss_surface", "surface losses"],
        "type": "float",
        "context_group": "mud_volume",
    },
    "mud_report.loss_downhole": {
        "keywords": ["lost down hole", "downhole loss", "loss downhole"],
        "synonyms": ["loss_downhole", "downhole losses", "lost downhole"],
        "type": "float",
        "context_group": "mud_volume",
    },
    # Drilling Params
    "drilling_params.bit_no": {
        "keywords": ["bit no", "bit number", "bit #"],
        "synonyms": ["bit_no", "bit num"],
        "type": "str",
        "context_group": "bit_info",
    },
    "drilling_params.bit_size": {
        "keywords": ["bit size", "hole size"],
        "synonyms": ["bit_size", "bit dia"],
        "type": "float",
        "context_group": "bit_info",
    },
    "drilling_params.bit_type": {
        "keywords": ["bit type", "iadc", "bit model"],
        "synonyms": ["bit_type", "bit make"],
        "type": "str",
        "context_group": "bit_info",
    },
    "drilling_params.avg_rop": {
        "keywords": ["avg. rop", "avg rop", "average rop", "rop"],
        "synonyms": ["avg_rop", "rate of penetration"],
        "type": "float",
        "valid_range": (0, 200),
        "context_group": "drilling_perf",
    },
    "drilling_params.depth_in": {
        "keywords": ["depth in", "bit depth in", "run in depth"],
        "synonyms": ["depth_in", "bit run in"],
        "type": "float",
        "valid_range": (0, 15000),
        "context_group": "bit_info",
    },
    "drilling_params.depth_out": {
        "keywords": ["depth out", "bit depth out", "run out depth"],
        "synonyms": ["depth_out", "bit run out"],
        "type": "float",
        "valid_range": (0, 15000),
        "context_group": "bit_info",
    },
    "drilling_params.hours_on_bottom": {
        "keywords": ["hours on bottom", "bit hours", "rotating hours"],
        "synonyms": ["hours_on_bottom", "bot hours"],
        "type": "float",
        "valid_range": (0, 500),
        "context_group": "drilling_perf",
    },
    "drilling_params.tfa": {
        "keywords": ["tfa", "total flow area"],
        "synonyms": ["t.f.a.", "flow area"],
        "type": "float",
        "valid_range": (0, 5),
        "context_group": "bit_info",
    },
    "drilling_params.wob_max": {
        "keywords": ["wob max", "wob", "weight on bit"],
        "synonyms": ["wob_max", "max wob"],
        "type": "float",
        "valid_range": (0, 100),
        "context_group": "drilling_params_block",
    },
    "drilling_params.rpm_max": {
        "keywords": ["rpm max", "rpm", "rotary speed"],
        "synonyms": ["rpm_max", "max rpm"],
        "type": "float",
        "valid_range": (0, 300),
        "context_group": "drilling_params_block",
    },
    "drilling_params.torque_max": {
        "keywords": ["torque max", "tq max", "torque"],
        "synonyms": ["torque_max", "max torque"],
        "type": "float",
        "valid_range": (0, 100),
        "context_group": "drilling_params_block",
    },
    "drilling_params.pump_pressure_max": {
        "keywords": ["spp max", "pump pressure max", "spp", "standpipe"],
        "synonyms": ["pump_pressure_max", "max spp", "stand pipe pressure"],
        "type": "float",
        "valid_range": (0, 8000),
        "context_group": "drilling_params_block",
    },
}

# Label indicators for rejecting false positives
LABEL_INDICATORS = {
    "field", "well", "client", "operator", "report", "mud", "bit",
    "depth", "rig", "drill", "supervisor", "geologist", "pusher",
    "manager", "section", "date", "type", "weight", "size", "shape",
    "water", "unit", "temperature", "viscosity", "loss", "filter",
    "gel", "fluid", "hardness", "chloride", "solid", "oil", "cake",
    "pump", "parameter", "output", "surface", "downhole", "cement",
    "casing", "from", "to", "hrs", "hours", "status", "code",
    "activity", "npt", "tool", "serial", "iadc", "drilled", "bottom",
    "equipment", "information", "data", "summary", "forecast",
    "description", "remark", "note", "total", "average", "max",
    "min", "pressure", "flow", "rate", "volume", "density",
}

# Field Labels for UI
FIELD_LABELS = {
    "well_info.name": "Well Name",
    "well_info.field_name": "Field Name",
    "well_info.well_type": "Well Type",
    "well_info.well_shape": "Well Shape",
    "well_info.client": "Client",
    "well_info.client_rep": "Client Rep",
    "well_info.operator": "Operator",
    "well_info.rig_name": "Rig Name",
    "well_info.drilling_contractor": "Drilling Contractor",
    "well_info.gle_msl": "GLE-MSL (m)",
    "well_info.rte_msl": "RTE-MSL (m)",
    "well_info.gle_rte": "GLE-RTE (m)",
    "well_info.target_depth": "Target Depth (m)",
    "well_info.water_depth": "Water Depth (m)",
    "well_info.latitude": "Latitude",
    "well_info.longitude": "Longitude",
    "well_info.northing": "Northing",
    "well_info.easting": "Easting",
    "well_info.formation": "Formation",
    "well_info.kop1": "KOP (m)",
    "well_info.rig_heading": "Rig Heading (°)",
    "well_info.lta_day": "LTA (Day)",
    "well_info.actual_rig_days": "Actual Rig Days",
    "well_info.section_name": "Hole Section",
    "well_info.spud_date": "Spud Date",
    "well_info.report_date": "Report Date",
    "well_info.start_hole_date": "Hole Section Start",
    "well_info.operation_manager": "Operation Manager",
    "well_info.superintendent": "Superintendent",
    "well_info.supervisor_day": "Supervisor (Day)",
    "well_info.supervisor_night": "Supervisor (Night)",
    "well_info.geologist1": "Geologist",
    "well_info.tool_pusher_day": "Tool Pusher",
    "daily_report.report_number": "Report Number",
    "daily_report.report_date": "Report Date",
    "daily_report.depth_0000": "Depth @ 00:00",
    "daily_report.depth_0600": "Depth @ 06:00",
    "daily_report.depth_2400": "Depth @ 24:00",
    "daily_report.rig_day": "Rig Day",
    "daily_report.summary": "Summary",
    "daily_report.forecast": "Forecast",
    "mud_report.mud_type": "Mud Type",
    "mud_report.mw": "Mud Weight (pcf)",
    "mud_report.pv": "PV (cp)",
    "mud_report.yp": "YP",
    "mud_report.funnel_vis": "Funnel Viscosity",
    "mud_report.gel_10s": "Gel 10s",
    "mud_report.gel_10m": "Gel 10m",
    "mud_report.fl": "Fluid Loss",
    "mud_report.ph": "pH",
    "mud_report.chloride": "Chloride",
    "mud_report.temperature": "Temperature",
    "mud_report.solid_percent": "Solids %",
    "mud_report.oil_percent": "Oil/Water %",
    "mud_report.cake_thickness": "Filter Cake",
    "mud_report.volume_hole": "Vol in Hole",
    "mud_report.loss_surface": "Loss Surface",
    "mud_report.loss_downhole": "Loss Downhole",
    "drilling_params.bit_no": "Bit No",
    "drilling_params.bit_size": "Bit Size (in)",
    "drilling_params.bit_type": "Bit Type",
    "drilling_params.avg_rop": "Avg ROP",
    "drilling_params.depth_in": "Depth In",
    "drilling_params.depth_out": "Depth Out",
    "drilling_params.wob_max": "WOB Max",
    "drilling_params.rpm_max": "RPM Max",
    "drilling_params.torque_max": "Torque Max",
    "drilling_params.pump_pressure_max": "SPP Max",
    "drilling_params.hours_on_bottom": "Hours on Bottom",
    "drilling_params.tfa": "TFA (in²)",
}

FIELD_GROUPS = {
    "--- Well Info ---": [k for k in FIELD_LABELS if k.startswith("well_info.")],
    "--- Daily Report ---": [k for k in FIELD_LABELS if k.startswith("daily_report.")],
    "--- Mud Report ---": [k for k in FIELD_LABELS if k.startswith("mud_report.")],
    "--- Drilling Params ---": [k for k in FIELD_LABELS if k.startswith("drilling_params.")],
}


# =====================================================
# Value Normalizer - centralized parsing
# =====================================================
class ValueNormalizer:
    """Central utility for parsing and normalizing values"""

    MONTH_MAP = {
        "jan": 1, "january": 1, "feb": 2, "february": 2,
        "mar": 3, "march": 3, "apr": 4, "april": 4,
        "may": 5, "jun": 6, "june": 6, "jul": 7, "july": 7,
        "aug": 8, "august": 8, "sep": 9, "september": 9,
        "oct": 10, "october": 10, "nov": 11, "november": 11,
        "dec": 12, "december": 12,
    }

    EMPTY_MARKERS = {"", "-", "--", "---", "n/a", "na", "none",
                     "null", "nan", "nil", "tbd", "tba"}

    @classmethod
    def is_empty(cls, value) -> bool:
        if value is None:
            return True
        s = str(value).strip().lower()
        return s in cls.EMPTY_MARKERS

    @classmethod
    def to_float(cls, value, valid_range=None) -> Optional[float]:
        if cls.is_empty(value):
            return None
        try:
            if isinstance(value, (int, float)):
                result = float(value)
            else:
                s = str(value).strip()
                s = re.sub(r'[,\s]', '', s)
                s = re.sub(r'[^\d\.\-eE]', '', s)
                if not s or s in ("-", ".", "-."):
                    return None
                result = float(s)
            if valid_range:
                lo, hi = valid_range
                if not (lo <= result <= hi):
                    return None
            if math.isnan(result) or math.isinf(result):
                return None
            return result
        except (ValueError, TypeError):
            return None

    @classmethod
    def to_int(cls, value) -> Optional[int]:
        if cls.is_empty(value):
            return None
        try:
            if isinstance(value, (int, float)):
                return int(value)
            s = re.sub(r'[^\d\-]', '', str(value).strip())
            if not s or s == "-":
                return None
            return int(float(s))
        except (ValueError, TypeError):
            return None

    @classmethod
    def to_date(cls, value) -> Optional[dt_date]:
        if cls.is_empty(value):
            return None
        if isinstance(value, dt_date):
            return value
        if isinstance(value, datetime):
            return value.date()
        s = str(value).strip()
        for fmt in (
            "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d",
            "%d-%m-%Y", "%m-%d-%Y", "%d.%m.%Y", "%Y.%m.%d",
            "%d %b %Y", "%d %B %Y", "%b %d %Y", "%B %d %Y",
            "%b %d, %Y", "%B %d, %Y",
        ):
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                continue
        try:
            num = float(s)
            if 1900 <= num <= 2100:
                return dt_date(int(num), 1, 1)
        except (ValueError, TypeError):
            pass
        return None

    @classmethod
    def to_time(cls, value) -> Optional[dt_time]:
        if value is None:
            return None
        if isinstance(value, dt_time):
            return value
        if isinstance(value, datetime):
            return value.time()
        s = str(value).strip()
        m = re.match(r'^(\d{1,2}):(\d{2})(?::(\d{2}))?', s)
        if m:
            h = int(m.group(1))
            mi = int(m.group(2))
            if h == 24:
                h = 0
            if 0 <= h <= 23 and 0 <= mi <= 59:
                return dt_time(h, mi)
        try:
            f = float(value)
            if 0 <= f < 1:
                total_sec = int(f * 86400)
                return dt_time(total_sec // 3600, (total_sec % 3600) // 60)
            elif 0 <= f <= 24:
                h = int(f)
                mi = int((f - h) * 60)
                if h == 24:
                    h = 0
                return dt_time(h, mi)
        except (ValueError, TypeError):
            pass
        return None

    @classmethod
    def to_str(cls, value) -> Optional[str]:
        if cls.is_empty(value):
            return None
        s = str(value).strip()
        return s if s else None

    @classmethod
    def convert(cls, value, data_type: str, valid_range=None):
        if data_type == "float":
            return cls.to_float(value, valid_range)
        elif data_type == "int":
            return cls.to_int(value)
        elif data_type == "date":
            return cls.to_date(value)
        elif data_type == "time":
            return cls.to_time(value)
        elif data_type == "text":
            return cls.to_str(value)
        else:
            return cls.to_str(value)

    @classmethod
    def combine_date_parts(cls, parts: list) -> Optional[dt_date]:
        cleaned = [str(v).strip() for v in parts
                   if v is not None and str(v).strip()]
        if not cleaned:
            return None
        if len(cleaned) == 1:
            return cls.to_date(cleaned[0])
        year = month = day = None
        for val in cleaned:
            vl = val.lower().strip(".")
            if vl in cls.MONTH_MAP or vl[:3] in cls.MONTH_MAP:
                month = cls.MONTH_MAP.get(vl, cls.MONTH_MAP.get(vl[:3]))
            else:
                try:
                    num = int(float(val))
                    if num > 1900:
                        year = num
                    elif month is not None and day is None:
                        day = num
                    elif month is None and 1 <= num <= 12:
                        month = num
                    elif day is None:
                        day = num
                    else:
                        year = num
                except (ValueError, TypeError):
                    continue
        if year and month and day:
            try:
                return dt_date(year, month, day)
            except ValueError:
                pass
        if year and month:
            try:
                return dt_date(year, month, 1)
            except ValueError:
                pass
        return None


# =====================================================
# Advanced Sheet Router
# =====================================================
class SheetRouter:
    """Smart sheet routing with content analysis"""

    def __init__(self, cell_cache: Dict[str, Dict]):
        self.cell_cache = cell_cache
        self.sheet_scores = {}

    def analyze(self) -> Dict[str, str]:
        self.sheet_scores = {}
        for sheet_name, cells in self.cell_cache.items():
            self.sheet_scores[sheet_name] = self._score_sheet(
                sheet_name, cells
            )

        routing = {}
        sections = [
            "well_info", "daily_report", "mud_report", "drilling_params"
        ]
        for section in sections:
            best_sheet, best_score = None, -1
            for sheet_name, scores in self.sheet_scores.items():
                score = scores.get(section, 0)
                if score > best_score:
                    best_score = score
                    best_sheet = sheet_name
            routing[section] = (
                best_sheet if best_sheet and best_score > 0
                else self._get_largest_sheet()
            )
        return routing

    def _score_sheet(self, name: str, cells: Dict) -> Dict[str, float]:
        scores = {
            "well_info": 0.0, "daily_report": 0.0,
            "mud_report": 0.0, "drilling_params": 0.0,
        }
        nl = name.lower()
        name_hints = {
            "well_info": [
                "remark", "ddr remark", "daily report",
                "report", "ddr", "header",
            ],
            "daily_report": [
                "remark", "ddr remark", "daily report",
                "report", "ddr", "time",
            ],
            "mud_report": [
                "data", "ddr data", "mud", "fluid",
                "drilling data",
            ],
            "drilling_params": [
                "data", "ddr data", "drilling", "bit",
                "parameter", "drilling data",
            ],
        }
        skip_hints = [
            "setting", "config", "template", "lookup",
            "reference", "list", "dropdown", "validation",
        ]
        for section, hints in name_hints.items():
            if any(h in nl for h in hints):
                scores[section] += 3.0
        if any(h in nl for h in skip_hints):
            for k in scores:
                scores[k] -= 10.0

        content_kw = {
            "well_info": [
                "well name", "client", "operator", "rig name",
                "spud", "field", "contractor",
            ],
            "daily_report": [
                "from", "hrs", "duration", "activity",
                "time", "operation",
            ],
            "mud_report": [
                "mud weight", "mw", "pv", "yp", "gel",
                "fluid loss", "funnel", "viscosity",
            ],
            "drilling_params": [
                "bit no", "bit size", "rop", "wob", "rpm",
                "torque", "spp", "tfa",
            ],
        }
        all_text = " ".join(
            str(v).lower() for v in cells.values() if v
        )
        for section, kws in content_kw.items():
            for kw in kws:
                if kw in all_text:
                    scores[section] += 0.5

        if "from" in all_text and "to" in all_text:
            if any(w in all_text for w in ["hrs", "duration", "hour"]):
                scores["daily_report"] += 3.0

        return scores

    def _get_largest_sheet(self) -> str:
        if not self.cell_cache:
            return ""
        return max(self.cell_cache, key=lambda k: len(self.cell_cache[k]))


# =====================================================
# Advanced Field Detector
# =====================================================
class FieldDetector:
    """
    Multi-strategy field detection engine:
    1. Exact keyword match
    2. Fuzzy matching with synonyms
    3. Context-group proximity scoring
    4. Radius-based value search
    """

    def __init__(self, cell_cache: Dict[str, Dict]):
        self.cell_cache = cell_cache
        self.found_locations = {}
        self.confidence_scores = {}

    def detect_field(
        self,
        field_path: str,
        pattern: dict,
        sheets_to_search: List[str],
    ) -> Optional[Tuple[Any, float, str, int, int]]:
        """
        Returns: (value, confidence, sheet_name, value_row, value_col)
        """
        keywords = pattern.get("keywords", [])
        synonyms = pattern.get("synonyms", [])
        data_type = pattern.get("type", "str")
        valid_range = pattern.get("valid_range")
        min_kw_len = pattern.get("min_keyword_length", 3)
        all_keywords = keywords + synonyms

        best = None
        best_conf = 0.0

        for sheet_name in sheets_to_search:
            cells = self.cell_cache.get(sheet_name, {})
            if not cells:
                continue

            for (row, col), cell_value in cells.items():
                cell_text = str(cell_value).strip()
                if len(cell_text) < min_kw_len or len(cell_text) > 120:
                    continue

                cell_lower = cell_text.lower()
                match_conf = self._best_keyword_match(
                    cell_lower, all_keywords
                )
                if match_conf < 0.55:
                    continue

                # context bonus
                context_bonus = self._context_bonus(
                    field_path, sheet_name, row, col
                )
                total_conf = min(1.0, match_conf + context_bonus)

                # extract value with radius search
                value_info = self._extract_value_radius(
                    sheet_name, row, col, data_type, valid_range
                )
                if not value_info:
                    continue

                value, v_row, v_col = value_info

                # reject if value looks like a label
                if self._is_label(str(value), field_path):
                    continue

                if total_conf > best_conf:
                    best_conf = total_conf
                    best = (
                        value, total_conf, sheet_name, v_row, v_col
                    )

        return best

    def _best_keyword_match(
        self, text: str, keywords: List[str]
    ) -> float:
        best = 0.0
        text = text.strip().rstrip(":")
        for kw in keywords:
            kw = kw.lower().strip()
            conf = self._match_score(text, kw)
            if conf > best:
                best = conf
        return best

    def _match_score(self, text: str, keyword: str) -> float:
        if text == keyword:
            return 1.0
        if text.rstrip(":") == keyword.rstrip(":"):
            return 0.98

        # short keywords need word boundary
        if len(keyword) <= 3:
            if re.search(rf"\b{re.escape(keyword)}\b", text):
                return 0.92
            return 0.0

        if text.startswith(keyword):
            return 0.93
        if keyword in text:
            ratio = len(keyword) / max(1, len(text))
            return 0.78 + ratio * 0.15

        # fuzzy
        seq_ratio = SequenceMatcher(None, text, keyword).ratio()
        if seq_ratio > 0.80:
            return seq_ratio * 0.85
        if seq_ratio > 0.70:
            return seq_ratio * 0.75

        # partial token match
        text_tokens = set(
            re.split(r'[\s/\-_\.\(\):]+', text)
        )
        kw_tokens = set(
            re.split(r'[\s/\-_\.\(\):]+', keyword)
        )
        if kw_tokens and text_tokens:
            overlap = len(kw_tokens & text_tokens) / len(kw_tokens)
            if overlap >= 0.7:
                return 0.6 + overlap * 0.2

        return 0.0

    def _context_bonus(
        self, field_path: str, sheet: str, row: int, col: int
    ) -> float:
        """Bonus if nearby cells contain related fields"""
        pattern = FIELD_PATTERNS.get(field_path, {})
        my_group = pattern.get("context_group", "")
        if not my_group:
            return 0.0

        bonus = 0.0
        for fp, loc in self.found_locations.items():
            if fp == field_path:
                continue
            other_pattern = FIELD_PATTERNS.get(fp, {})
            other_group = other_pattern.get("context_group", "")
            if other_group != my_group:
                continue
            if loc["sheet"] != sheet:
                continue

            dist = abs(loc["row"] - row) + abs(loc["col"] - col)
            if dist <= 5:
                bonus += 0.08
            elif dist <= 15:
                bonus += 0.04
            elif dist <= 30:
                bonus += 0.02

        return min(bonus, 0.15)

    def _extract_value_radius(
        self,
        sheet_name: str,
        label_row: int,
        label_col: int,
        data_type: str,
        valid_range=None,
    ) -> Optional[Tuple[Any, int, int]]:
        """
        Radius-based value extraction:
        Search in expanding radius around label
        """
        cells = self.cell_cache.get(sheet_name, {})
        if not cells:
            return None

        candidates = []

        # Strategy 1: Right (same row, cols +1 to +15)
        for dc in range(1, 16):
            c = label_col + dc
            val = cells.get((label_row, c))
            if val is None:
                continue
            if ValueNormalizer.is_empty(val):
                continue
            if self._is_label(str(val)):
                continue
            converted = ValueNormalizer.convert(
                val, data_type, valid_range
            )
            if converted is not None:
                candidates.append((converted, label_row, c, 1.0 - dc * 0.03))

        # Strategy 2: Below (rows +1 to +3, cols 0 to +5)
        for dr in range(1, 4):
            for dc in range(0, 6):
                r = label_row + dr
                c = label_col + dc
                val = cells.get((r, c))
                if val is None:
                    continue
                if ValueNormalizer.is_empty(val):
                    continue
                if self._is_label(str(val)):
                    continue
                converted = ValueNormalizer.convert(
                    val, data_type, valid_range
                )
                if converted is not None:
                    score = 0.85 - dr * 0.1 - dc * 0.03
                    candidates.append((converted, r, c, score))

        # Strategy 3: Left (for right-to-left layouts)
        for dc in range(1, 4):
            c = label_col - dc
            if c < 1:
                break
            val = cells.get((label_row, c))
            if val is None:
                continue
            if ValueNormalizer.is_empty(val):
                continue
            if self._is_label(str(val)):
                continue
            converted = ValueNormalizer.convert(
                val, data_type, valid_range
            )
            if converted is not None:
                candidates.append((converted, label_row, c, 0.6 - dc * 0.1))

        if not candidates:
            return None

        candidates.sort(key=lambda x: x[3], reverse=True)
        best = candidates[0]
        return (best[0], best[1], best[2])

    def _is_label(
        self, text: str, field_path: str = ""
    ) -> bool:
        s = text.strip()
        s_lower = s.lower()

        if not s:
            return True

        # numeric -> not a label
        try:
            float(s.replace(",", "").replace("+", ""))
            return False
        except (ValueError, TypeError):
            pass

        # date-like -> not a label
        if ValueNormalizer.to_date(s) is not None:
            return False

        # short alpha -> not a label (likely abbreviation/value)
        if len(s) <= 3 and s.isalpha():
            return False

        # month names -> not a label
        month_names = [
            "jan", "feb", "mar", "apr", "may", "jun",
            "jul", "aug", "sep", "oct", "nov", "dec",
        ]
        if any(m in s_lower for m in month_names):
            return False

        # ends with colon -> label
        if s.endswith(":"):
            return True

        # too long -> label
        if len(s) > 50:
            return True

        # check if the value matches this field's own keywords
        if field_path:
            pattern = FIELD_PATTERNS.get(field_path, {})
            all_kw = (
                pattern.get("keywords", []) +
                pattern.get("synonyms", [])
            )
            for kw in all_kw:
                if s_lower.rstrip(":") == kw.lower().rstrip(":"):
                    return True

        # high ratio of label indicator words
        words = set(re.split(r'[\s/\-_\.\(\)]+', s_lower))
        words = {w for w in words if len(w) >= 2}
        if words:
            indicator_count = sum(
                1 for w in words if w in LABEL_INDICATORS
            )
            ratio = indicator_count / len(words)
            if ratio >= 0.5 and len(words) >= 2:
                return True

        return False

    def register_found(
        self,
        field_path: str,
        sheet: str,
        row: int,
        col: int,
        confidence: float,
    ):
        self.found_locations[field_path] = {
            "sheet": sheet,
            "row": row,
            "col": col,
            "confidence": confidence,
        }
        self.confidence_scores[field_path] = confidence


# =====================================================
# Code Resolver - full name resolution
# =====================================================
class CodeResolver:
    """Resolve main codes and sub codes to full names"""

    @staticmethod
    def _clean_code(value):
        """Normalize Excel numeric/text codes without losing decimals."""
        if value is None:
            return ""
        text = str(value).strip().replace("–", "-").replace("—", "-")
        # Excel commonly turns integer codes into 2.0; only remove a
        # trailing numeric .0, never a meaningful composite such as 2.1.
        text = re.sub(r"(?<=\d)\.0+(?=\s*$)", "", text)
        return text

    @staticmethod
    def _main_number(value):
        text = CodeResolver._clean_code(value)
        # Accept 2.1, 2-Drilling, 2 / Drilling and "2 - Drilling".
        match = re.search(r"^\s*(\d+)", text)
        return match.group(1) if match else ""

    @staticmethod
    def _composite_number(value):
        text = CodeResolver._clean_code(value)
        match = re.search(r"^\s*(\d+)\s*[./-]\s*(\d+)", text)
        return f"{match.group(1)}.{match.group(2)}" if match else ""

    @staticmethod
    def resolve_main_code(raw_code) -> str:
        if raw_code is None:
            return ""
        code_str = CodeResolver._clean_code(raw_code)
        composite = CodeResolver._composite_number(code_str)
        clean = CodeResolver._main_number(composite or code_str)

        # already has name
        if " - " in code_str and len(code_str) > 5:
            return code_str

        # try direct lookup
        if clean in MAIN_CODE_MAP:
            return f"{clean} - {MAIN_CODE_MAP[clean]}"

        # try as int
        try:
            int_code = str(int(float(clean)))
            if int_code in MAIN_CODE_MAP:
                return f"{int_code} - {MAIN_CODE_MAP[int_code]}"
        except (ValueError, TypeError):
            pass

        # try reverse lookup (name -> code)
        lower = code_str.lower().strip()
        if lower in MAIN_CODE_REVERSE:
            code_num = MAIN_CODE_REVERSE[lower]
            return f"{code_num} - {MAIN_CODE_MAP[code_num]}"

        # fuzzy match on name
        best_match, best_ratio = None, 0.0
        for name, code_num in MAIN_CODE_REVERSE.items():
            ratio = SequenceMatcher(None, lower, name).ratio()
            if ratio > best_ratio and ratio > 0.7:
                best_ratio = ratio
                best_match = code_num
        if best_match:
            return f"{best_match} - {MAIN_CODE_MAP[best_match]}"

        return code_str

    @staticmethod
    def resolve_sub_code(raw_sub, raw_main="") -> str:
        # A surprising number of DDR sheets put the composite code (2.3) in
        # either the main or sub column. Prefer the explicit composite.
        composite_from_sub = CodeResolver._composite_number(raw_sub)
        composite_from_main = CodeResolver._composite_number(raw_main)
        composite = composite_from_sub or composite_from_main
        if composite and composite in SUB_CODE_MAP:
            return f"{composite} - {SUB_CODE_MAP[composite]}"
        if raw_sub is None or str(raw_sub).strip() == "":
            return ""
        sub_str = CodeResolver._clean_code(raw_sub)
        clean_sub = sub_str

        # already has name
        if len(sub_str) > 8 and any(
            c.isalpha() for c in sub_str
        ):
            return sub_str

        # get main code number
        main_num = ""
        if raw_main:
            main_str = str(raw_main).strip()
            try:
                main_num = str(int(float(
                    main_str.split("-")[0].split(" ")[0]
                    .replace(".0", "").strip()
                )))
            except (ValueError, TypeError):
                pass

        # try sub code number
        try:
            sub_num = str(int(float(clean_sub)))
        except (ValueError, TypeError):
            sub_num = clean_sub

        # composite key: main.sub
        if main_num:
            composite = f"{main_num}.{sub_num}"
            if composite in SUB_CODE_MAP:
                return f"{composite} - {SUB_CODE_MAP[composite]}"

        # try direct lookup
        if clean_sub in SUB_CODE_MAP:
            return f"{clean_sub} - {SUB_CODE_MAP[clean_sub]}"

        # try all composites with this sub number
        for key, name in SUB_CODE_MAP.items():
            if key.endswith(f".{sub_num}"):
                if main_num and key.startswith(f"{main_num}."):
                    return f"{key} - {name}"

        # fuzzy on name
        lower = sub_str.lower().strip()
        if lower in SUB_CODE_REVERSE:
            code = SUB_CODE_REVERSE[lower]
            return f"{code} - {SUB_CODE_MAP[code]}"

        best_match, best_ratio = None, 0.0
        for name, code in SUB_CODE_REVERSE.items():
            ratio = SequenceMatcher(None, lower, name).ratio()
            if ratio > best_ratio and ratio > 0.7:
                best_ratio = ratio
                best_match = code
        if best_match:
            return f"{best_match} - {SUB_CODE_MAP[best_match]}"

        return sub_str

    @staticmethod
    def guess_contractor(npt_code: str) -> str:
        if not npt_code:
            return ""
        code_lower = npt_code.lower().strip()

        # exact match
        if code_lower in NPT_CONTRACTOR_MAP:
            return NPT_CONTRACTOR_MAP[code_lower]

        # prefix match
        for prefix, contractor in NPT_CONTRACTOR_MAP.items():
            if code_lower.startswith(prefix):
                return contractor

        # category guess
        if code_lower.startswith("rr"):
            return "Rig Contractor"
        if code_lower.startswith("f-"):
            return "Service Company"
        if code_lower.startswith("w-"):
            return "Operations"
        if code_lower.startswith("t-"):
            return "Operations"

        return ""


# =====================================================
# Learning Manager - learns from user corrections
# =====================================================
class LearningManager:
    """Stores and applies user correction patterns"""

    def __init__(self):
        self.corrections = {}
        self._load()

    def _get_path(self) -> str:
        d = os.path.join(
            os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
            "templates",
        )
        os.makedirs(d, exist_ok=True)
        return os.path.join(d, "_learning_data.json")

    def _load(self):
        path = self._get_path()
        if os.path.exists(path):
            try:
                with open(path, "r", encoding="utf-8") as f:
                    self.corrections = json.load(f)
            except Exception:
                self.corrections = {}

    def save(self):
        try:
            with open(self._get_path(), "w", encoding="utf-8") as f:
                json.dump(
                    self.corrections, f, indent=2,
                    ensure_ascii=False, default=str,
                )
        except Exception as e:
            logger.error(f"Learning save error: {e}")

    def record_correction(
        self,
        field_path: str,
        anchor_text: str,
        direction: str,
        offset_row: int,
        offset_col: int,
    ):
        if field_path not in self.corrections:
            self.corrections[field_path] = []

        entry = {
            "anchor": anchor_text.lower().strip(),
            "direction": direction,
            "offset_row": offset_row,
            "offset_col": offset_col,
            "count": 1,
            "last_used": datetime.now().isoformat(),
        }

        # update existing or add new
        for existing in self.corrections[field_path]:
            if existing["anchor"] == entry["anchor"]:
                existing["count"] += 1
                existing["last_used"] = entry["last_used"]
                existing["direction"] = direction
                existing["offset_row"] = offset_row
                existing["offset_col"] = offset_col
                break
        else:
            self.corrections[field_path].append(entry)

        self.save()

    def get_learned_patterns(
        self, field_path: str
    ) -> List[Dict]:
        patterns = self.corrections.get(field_path, [])
        return sorted(
            patterns,
            key=lambda x: x.get("count", 0),
            reverse=True,
        )


# =====================================================
# Main Dialog
# =====================================================
class SmartTemplateDialog(QDialog):
    """
    Unified Import Dialog v2.0:
    - Smart Engine with multi-strategy detection
    - Anchor-based templates
    - Learning from corrections
    - Full code resolution
    - Context-aware scoring
    """

    import_completed = Signal(dict)

    def __init__(
        self,
        db_manager,
        well_id: int,
        parent=None,
        preload_template: dict = None,
        preload_file: str = "",
    ):
        super().__init__(parent)
        self.db = db_manager
        self.well_id = well_id

        self.wb = None
        self.cell_cache = {}
        self.assignments = {}
        self.confidence_scores = {}
        self.sheet_routing = {}
        self.sheet_checkboxes = {}
        self._pending_cell = None
        self._selected_field = None
        self._range_mode = False
        self._range_cells = []
        self.filepath = preload_file
        self.preload_template = preload_template or {}
        self.current_sheet = ""

        self.base_extracted = {}
        self.final_data = {}

        self.detector = None
        self.learning = LearningManager()

        self.setWindowTitle("🧠 Smart Template Import v2.0")
        self.setMinimumSize(1300, 800)
        self.setModal(True)
        self._init_ui()

        if self.filepath:
            self.file_label.setText(self.filepath)
            self._load_workbook(self.filepath)
        if self.preload_template:
            self._apply_template(self.preload_template)

    # ================================================================
    # UI
    # ================================================================
    def _init_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(8, 8, 8, 8)
        layout.setSpacing(6)

        # Header
        header = QLabel(
            "🧠 Smart Template Import v2.0 — "
            "Auto-Detect + Anchor Templates + Learning"
        )
        header.setStyleSheet(
            "font-size: 13px; font-weight: bold; color: #2c3e50; "
            "padding: 6px 10px; background: #ecf0f1; border-radius: 4px;"
        )
        header.setFixedHeight(30)
        layout.addWidget(header)

        # File row
        file_layout = QHBoxLayout()
        self.file_label = QLineEdit()
        self.file_label.setReadOnly(True)
        self.file_label.setPlaceholderText("Select Excel file...")
        self.file_label.setFixedHeight(28)
        browse_btn = QPushButton("📂 Open Excel")
        browse_btn.setFixedHeight(28)
        browse_btn.setStyleSheet(
            "background: #3498db; color: white; padding: 4px 12px; "
            "border-radius: 4px; font-weight: bold; border: none;"
        )
        browse_btn.clicked.connect(self._browse_file)
        file_layout.addWidget(self.file_label)
        file_layout.addWidget(browse_btn)
        layout.addLayout(file_layout)

        # Sheet bar
        sheet_bar = QWidget()
        sheet_bar.setStyleSheet(
            "background: #f8f9fa; border-radius: 3px;"
        )
        sheet_layout = QHBoxLayout(sheet_bar)
        sheet_layout.setContentsMargins(6, 4, 6, 4)
        sheet_layout.addWidget(QLabel("📄 Sheets:"))
        self.sheet_checkboxes_widget = QWidget()
        self.sheet_checkboxes_layout = QHBoxLayout(
            self.sheet_checkboxes_widget
        )
        self.sheet_checkboxes_layout.setContentsMargins(0, 0, 0, 0)
        self.sheet_checkboxes_layout.setSpacing(6)
        sheet_layout.addWidget(self.sheet_checkboxes_widget)
        sheet_layout.addStretch()
        sheet_layout.addWidget(QLabel("View:"))
        self.sheet_combo = QComboBox()
        self.sheet_combo.setMinimumWidth(140)
        self.sheet_combo.currentIndexChanged.connect(
            self._on_sheet_changed
        )
        sheet_layout.addWidget(self.sheet_combo)
        sheet_bar.setFixedHeight(32)
        layout.addWidget(sheet_bar)

        # Action bar
        auto_layout = QHBoxLayout()
        self.smart_btn = QPushButton("🧠 Smart Auto-Detect")
        self.smart_btn.setStyleSheet(
            "background: #e67e22; color: white; padding: 8px 16px; "
            "border-radius: 4px; font-weight: bold; border: none;"
        )
        self.smart_btn.setEnabled(False)
        self.smart_btn.clicked.connect(self._smart_auto_detect)
        auto_layout.addWidget(self.smart_btn)

        self.detect_status = QLabel("")
        self.detect_status.setStyleSheet(
            "color: #666; font-size: 10px;"
        )
        auto_layout.addWidget(self.detect_status)
        auto_layout.addStretch()

        self.load_tmpl_btn = QPushButton("📂 Load Template")
        self.load_tmpl_btn.setStyleSheet(
            "background: #1abc9c; color: white; padding: 8px 12px; "
            "border-radius: 4px; font-weight: bold; border: none;"
        )
        self.load_tmpl_btn.clicked.connect(self._load_template_file)
        auto_layout.addWidget(self.load_tmpl_btn)
        layout.addLayout(auto_layout)

        # Main splitter
        splitter = QSplitter(Qt.Horizontal)

        # LEFT: Review
        left = QWidget()
        left_layout = QVBoxLayout(left)
        left_layout.setContentsMargins(0, 0, 0, 0)
        left_layout.setSpacing(4)

        self.summary_label = QLabel("📊 No data detected yet")
        self.summary_label.setStyleSheet(
            "font-size: 10px; font-weight: bold; padding: 4px; "
            "background: #f0f0f0; border-radius: 3px; color: #555;"
        )
        self.summary_label.setFixedHeight(22)
        left_layout.addWidget(self.summary_label)
        self.code_detection_label = QLabel("🏷️ Activity codes: not scanned")
        self.code_detection_label.setStyleSheet("font-size: 10px; color: #666; padding: 3px;")
        self.code_detection_label.setWordWrap(True)
        left_layout.addWidget(self.code_detection_label)

        fl = QHBoxLayout()
        fl.addWidget(QLabel("Filter:"))
        self.filter_combo = QComboBox()
        self.filter_combo.setFixedHeight(24)
        self.filter_combo.addItems([
            "All", "🟢 High Conf.", "🟡 Medium", "🔴 Missing",
            "Well", "Report", "Mud", "Drilling",
        ])
        self.filter_combo.currentIndexChanged.connect(
            self._filter_review
        )
        fl.addWidget(self.filter_combo)
        fl.addStretch()
        left_layout.addLayout(fl)

        self.review_table = QTableWidget(0, 4)
        self.review_table.setHorizontalHeaderLabels([
            "", "Field", "Value", "Conf",
        ])
        self.review_table.setColumnWidth(0, 30)
        self.review_table.setColumnWidth(1, 150)
        self.review_table.setColumnWidth(3, 50)
        self.review_table.horizontalHeader().setSectionResizeMode(
            2, QHeaderView.Stretch
        )
        self.review_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.review_table.setAlternatingRowColors(True)
        self.review_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.review_table.verticalHeader().setDefaultSectionSize(22)
        self.review_table.verticalHeader().setVisible(False)
        self.review_table.itemClicked.connect(self._on_review_clicked)
        self.review_table.setStyleSheet("""
            QTableWidget { font-size: 10px; gridline-color: #eee; }
            QTableWidget::item:selected {
                background: #3498db; color: white;
            }
            QHeaderView::section {
                background: #34495e; color: white;
                padding: 3px; font-size: 9px; font-weight: bold;
            }
        """)
        left_layout.addWidget(self.review_table)
        splitter.addWidget(left)

        # RIGHT: Excel preview
        right = QWidget()
        right_layout = QVBoxLayout(right)
        right_layout.setContentsMargins(0, 0, 0, 0)
        right_layout.setSpacing(4)

        self.excel_table = QTableWidget()
        self.excel_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.excel_table.setSelectionMode(QTableWidget.SingleSelection)
        self.excel_table.cellClicked.connect(self._on_excel_clicked)
        self.excel_table.setStyleSheet("""
            QTableWidget { font-size: 10px; gridline-color: #ddd; }
            QTableWidget::item:selected {
                background: #e74c3c; color: white;
            }
            QHeaderView::section {
                background: #2c3e50; color: white;
                font-size: 9px; padding: 3px;
            }
        """)
        self.excel_table.verticalHeader().setDefaultSectionSize(20)
        right_layout.addWidget(self.excel_table)

        # Assign bar
        assign_bar = QHBoxLayout()
        assign_bar.setSpacing(6)
        self.cell_info = QLabel(
            "← Click a field (left), then click cell (right)"
        )
        self.cell_info.setStyleSheet(
            "color: #666; font-size: 10px; padding: 4px; "
            "background: #f8f9fa; border: 1px solid #dee2e6; "
            "border-radius: 3px;"
        )
        self.cell_info.setWordWrap(True)
        self.cell_info.setFixedHeight(44)
        assign_bar.addWidget(self.cell_info, 1)

        self.assign_btn = QPushButton("✅ Assign")
        self.assign_btn.setFixedSize(80, 28)
        self.assign_btn.setStyleSheet(
            "background: #27ae60; color: white; font-weight: bold; "
            "border-radius: 3px; border: none;"
        )
        self.assign_btn.setEnabled(False)
        self.assign_btn.clicked.connect(self._assign_cell)
        assign_bar.addWidget(self.assign_btn)
        right_layout.addLayout(assign_bar)

        # Range mode
        range_layout = QHBoxLayout()
        self.range_btn = QPushButton("📐 Range Mode: OFF")
        self.range_btn.setCheckable(True)
        self.range_btn.setFixedHeight(24)
        self.range_btn.setStyleSheet(
            "background: #7f8c8d; color: white; font-weight: bold; "
            "font-size: 9px; border-radius: 3px; border: none;"
        )
        self.range_btn.toggled.connect(self._toggle_range_mode)
        range_layout.addWidget(self.range_btn)
        range_layout.addStretch()
        right_layout.addLayout(range_layout)

        splitter.addWidget(right)
        splitter.setSizes([420, 880])
        layout.addWidget(splitter)

        # Bottom bar
        bottom = QHBoxLayout()
        bottom.addWidget(QLabel("Template:"))
        self.template_name = QLineEdit()
        self.template_name.setPlaceholderText(
            "template name (optional)"
        )
        self.template_name.setFixedWidth(200)
        bottom.addWidget(self.template_name)

        self.save_tmpl_btn = QPushButton("💾 Save Template")
        self.save_tmpl_btn.setStyleSheet(
            "background: #9b59b6; color: white; padding: 8px 16px; "
            "font-weight: bold; border-radius: 4px; border: none;"
        )
        self.save_tmpl_btn.clicked.connect(self._save_template)
        bottom.addWidget(self.save_tmpl_btn)
        bottom.addStretch()

        cancel_btn = QPushButton("Cancel")
        cancel_btn.clicked.connect(self.reject)
        bottom.addWidget(cancel_btn)

        self.import_btn = QPushButton("🚀 Import to Database")
        self.import_btn.setStyleSheet(
            "background: #27ae60; color: white; padding: 10px 20px; "
            "font-size: 12px; font-weight: bold; "
            "border-radius: 4px; border: none;"
        )
        self.import_btn.setEnabled(False)
        self.import_btn.clicked.connect(self._do_import)
        bottom.addWidget(self.import_btn)
        layout.addLayout(bottom)

    # ================================================================
    # File & Cache
    # ================================================================
    def _browse_file(self):
        filepath, _ = QFileDialog.getOpenFileName(
            self, "Open Excel File", "",
            "Excel Files (*.xlsx *.xls)",
        )
        if filepath:
            self.filepath = filepath
            self.file_label.setText(filepath)
            self._load_workbook(filepath)

    def _load_workbook(self, filepath: str):
        try:
            self.wb = load_workbook(filepath, data_only=True)
            self.cell_cache.clear()
            self.assignments.clear()
            self.confidence_scores.clear()
            self.sheet_routing.clear()

            for ws in self.wb.worksheets:
                self.cell_cache[ws.title] = self._unmerge(ws)

            self.sheet_combo.blockSignals(True)
            self.sheet_combo.clear()

            while self.sheet_checkboxes_layout.count():
                item = self.sheet_checkboxes_layout.takeAt(0)
                if item.widget():
                    item.widget().deleteLater()
            self.sheet_checkboxes.clear()

            for ws in self.wb.worksheets:
                name = ws.title
                self.sheet_combo.addItem(name)
                cb = QCheckBox(name)
                cb.setChecked(True)
                cb.setStyleSheet("font-size: 10px;")
                self.sheet_checkboxes[name] = cb
                self.sheet_checkboxes_layout.addWidget(cb)

            self.sheet_combo.blockSignals(False)

            if self.wb.worksheets:
                self.current_sheet = self.wb.worksheets[0].title
                self._display_sheet(self.current_sheet)
                self.smart_btn.setEnabled(True)
                self._reset_review_table()

            if self.wb and self.wb.worksheets and not self.preload_template:
                QTimer.singleShot(300, self._smart_auto_detect)

        except Exception as e:
            logger.error(f"Load error: {e}", exc_info=True)
            QMessageBox.critical(
                self, "Error", f"Cannot open:\n{e}"
            )

    def _unmerge(self, ws) -> Dict:
        """
        Read worksheet cells with proper merged cell handling.
        Only keep the master cell value, skip slave cells.
        """
        cells = {}
        merged_slaves = set()

        try:
            for mr in ws.merged_cells.ranges:
                min_c, min_r, max_c, max_r = mr.bounds
                for r in range(min_r, max_r + 1):
                    for c in range(min_c, max_c + 1):
                        if r != min_r or c != min_c:
                            merged_slaves.add((r, c))
        except Exception:
            pass

        for r in range(1, min(ws.max_row + 1, MAX_SCAN_ROWS)):
            for c in range(1, min(ws.max_column + 1, MAX_SCAN_COLS)):
                if (r, c) in merged_slaves:
                    continue
                val = ws.cell(row=r, column=c).value
                if val is not None:
                    cells[(r, c)] = val

        return cells

    def _display_sheet(self, sheet_name: str):
        self.excel_table.setRowCount(0)
        self.excel_table.setColumnCount(0)
        cells = self.cell_cache.get(sheet_name, {})
        if not cells:
            return

        max_r = min(
            max((r for r, c in cells.keys()), default=1),
            MAX_PREVIEW_ROWS,
        )
        max_c = min(
            max((c for r, c in cells.keys()), default=1),
            MAX_PREVIEW_COLS,
        )

        self.excel_table.setRowCount(max_r)
        self.excel_table.setColumnCount(max_c)
        headers = [get_column_letter(c) for c in range(1, max_c + 1)]
        self.excel_table.setHorizontalHeaderLabels(headers)

        for (r, c), val in cells.items():
            if r > max_r or c > max_c:
                continue
            text = str(val)[:100]
            item = QTableWidgetItem(text)
            item.setToolTip(str(val))

            # highlight assigned cells
            for fp, assign in self.assignments.items():
                if (assign.get("sheet") == sheet_name
                    and assign.get("row") == r
                    and assign.get("col") == c):
                    conf = assign.get("confidence", 1.0)
                    if conf >= 0.8:
                        item.setBackground(QColor("#d5f5e3"))
                    elif conf >= 0.6:
                        item.setBackground(QColor("#fef9e7"))
                    else:
                        item.setBackground(QColor("#d6eaf8"))
                    item.setToolTip(f"✅ {fp}\n{str(val)}")
                    break

            self.excel_table.setItem(r - 1, c - 1, item)

        self.excel_table.resizeColumnsToContents()
        for c in range(self.excel_table.columnCount()):
            w = self.excel_table.columnWidth(c)
            self.excel_table.setColumnWidth(c, max(40, min(w, 180)))

    def _on_sheet_changed(self, index):
        if index >= 0:
            self.current_sheet = self.sheet_combo.currentText()
            self._display_sheet(self.current_sheet)

    # ================================================================
    # Smart Auto-Detect v2
    # ================================================================
    def _smart_auto_detect(self):
        if not self.cell_cache:
            return

        self.detect_status.setText("🧠 Smart detecting...")
        QApplication.processEvents()

        try:
            # Route sheets
            router = SheetRouter(self.cell_cache)
            self.sheet_routing = router.analyze()

            # Initialize
            self.base_extracted = {
                "well_info": {}, "daily_report": {},
                "mud_report": {}, "drilling_params": {},
                "time_logs_24h": [], "time_logs_morning": [],
                "metadata": {"sheet_routing": self.sheet_routing},
            }
            self.assignments.clear()
            self.confidence_scores.clear()

            # Create detector
            self.detector = FieldDetector(self.cell_cache)

            # Sort fields by priority (high priority first)
            sorted_fields = sorted(
                FIELD_PATTERNS.items(),
                key=lambda x: x[1].get("priority", 5),
                reverse=True,
            )

            # Phase 1: detect all fields
            for field_path, pattern in sorted_fields:
                section = field_path.split(".", 1)[0]
                priority_sheet = self.sheet_routing.get(section)

                sheets_to_search = []
                if priority_sheet and priority_sheet in self.cell_cache:
                    sheets_to_search.append(priority_sheet)
                for sn in self.cell_cache:
                    if sn not in sheets_to_search:
                        score = router.sheet_scores.get(
                            sn, {}
                        ).get(section, 0)
                        if score > -5:
                            sheets_to_search.append(sn)

                # Try learned patterns first
                learned = self.learning.get_learned_patterns(field_path)
                if learned:
                    result = self._try_learned_patterns(
                        field_path, pattern, learned
                    )
                    if result:
                        self._register_detection(
                            field_path, section, *result
                        )
                        continue

                # Smart detect
                result = self.detector.detect_field(
                    field_path, pattern, sheets_to_search
                )
                if result:
                    value, conf, sheet, v_row, v_col = result
                    if not self._should_reject(
                        field_path, value, sheet
                    ):
                        self._register_detection(
                            field_path, section,
                            value, conf, sheet, v_row, v_col,
                        )

            # Phase 2: Text blocks (summary, forecast)
            self._detect_text_blocks()

            # Phase 3: Time logs
            self._detect_time_logs()

            # Build final data
            self.final_data = self._build_final_data_from_assignments()
            self._reset_review_table()
            self._display_sheet(self.current_sheet)

            detected = len(self.assignments)
            self.detect_status.setText(
                f"✅ Detected {detected} fields "
                f"(sheets: {len(self.sheet_routing)})"
            )

        except Exception as e:
            logger.error(
                f"Smart detect error: {e}", exc_info=True
            )
            self.detect_status.setText("❌ Detection failed")

    def _register_detection(
        self,
        field_path: str,
        section: str,
        value,
        conf: float,
        sheet: str,
        v_row: int,
        v_col: int,
    ):
        """Register a detected field"""
        key = field_path.split(".", 1)[1]
        self.assignments[field_path] = {
            "sheet": sheet,
            "row": v_row,
            "col": v_col,
            "value": str(value)[:100] if value is not None else "",
            "confidence": conf,
            "auto": True,
        }
        self.confidence_scores[field_path] = conf
        self.base_extracted[section][key] = value

        if self.detector:
            self.detector.register_found(
                field_path, sheet, v_row, v_col, conf
            )

    def _try_learned_patterns(
        self,
        field_path: str,
        pattern: dict,
        learned: List[Dict],
    ) -> Optional[Tuple]:
        """Try to find field using learned patterns"""
        data_type = pattern.get("type", "str")
        valid_range = pattern.get("valid_range")

        for lp in learned[:3]:
            anchor = lp["anchor"]
            for sheet_name, cells in self.cell_cache.items():
                for (r, c), val in cells.items():
                    cell_text = str(val).strip().lower()
                    if anchor in cell_text or cell_text in anchor:
                        # found anchor
                        vr = r + lp.get("offset_row", 0)
                        vc = c + lp.get("offset_col", 1)
                        cell_val = cells.get((vr, vc))
                        if cell_val is not None:
                            converted = ValueNormalizer.convert(
                                cell_val, data_type, valid_range
                            )
                            if converted is not None:
                                return (
                                    converted, 0.90,
                                    sheet_name, vr, vc,
                                )
        return None

    def _should_reject(
        self, field_path: str, value, sheet_name: str
    ) -> bool:
        if field_path == "well_info.section_name":
            if isinstance(value, (int, float)):
                return True
            val = str(value).strip().lower()
            if re.fullmatch(r"\d{4}", val):
                return True
            if val in [
                "jan", "feb", "mar", "company", "project",
                "status", "name", "type",
            ]:
                return True

        if field_path == "mud_report.mw":
            if "setting" in sheet_name.lower():
                return True

        return False

    def _detect_text_blocks(self):
        """Detect summary and forecast text blocks"""
        daily_sheet = self.sheet_routing.get("daily_report")
        sheets = []
        if daily_sheet:
            sheets.append(daily_sheet)
        for sn in self.cell_cache:
            if sn not in sheets:
                sheets.append(sn)

        for sn in sheets:
            if "summary" not in self.base_extracted.get(
                "daily_report", {}
            ):
                txt = self._find_text_block(
                    sn,
                    [
                        "summary of activities",
                        "summary of operation",
                        "remarks", "operations summary",
                    ],
                    [
                        "operation forecast", "forecast",
                        "from", "prepared by",
                    ],
                )
                if txt and len(txt) > 10:
                    self.base_extracted["daily_report"]["summary"] = txt

        for sn in sheets:
            if "forecast" not in self.base_extracted.get(
                "daily_report", {}
            ):
                txt = self._find_text_block(
                    sn,
                    [
                        "operation forecast", "forecast for next",
                        "planned operations", "next 24",
                    ],
                    ["from", "time", "prepared by", "approved"],
                )
                if txt and len(txt) > 10:
                    self.base_extracted["daily_report"]["forecast"] = txt

    def _find_text_block(
        self,
        sheet_name: str,
        start_keywords: List[str],
        stop_keywords: List[str],
    ) -> str:
        cells = self.cell_cache.get(sheet_name, {})
        if not cells:
            return ""

        start_row = None
        for (r, c), val in sorted(cells.items()):
            for kw in start_keywords:
                if kw.lower() in str(val).lower():
                    start_row = r
                    break
            if start_row:
                break

        if start_row is None:
            return ""

        lines = []
        for r in range(start_row + 1, start_row + 30):
            row_vals = {
                c: v for (r2, c), v in cells.items() if r2 == r
            }
            if not row_vals:
                continue
            row_text = " ".join(
                str(v).strip()
                for c, v in sorted(row_vals.items())
                if str(v).strip()
            ).strip()
            if not row_text:
                continue
            if any(
                sk.lower() in row_text.lower()
                for sk in stop_keywords
            ):
                break
            if len(row_text) > 3:
                lines.append(row_text)

        return "\n".join(lines)

    def _detect_time_logs(self):
        """Detect all time log tables"""
        daily_sheet = self.sheet_routing.get("daily_report")
        sheets = []
        if daily_sheet:
            sheets.append(daily_sheet)
        for sn in self.cell_cache:
            if sn not in sheets:
                sheets.append(sn)

        for sn in sheets:
            if (self.base_extracted["time_logs_24h"]
                and self.base_extracted["time_logs_morning"]):
                break
            tables = self._find_all_time_tables(sn)
            if tables:
                if len(tables) >= 1 and not self.base_extracted[
                    "time_logs_24h"
                ]:
                    self.base_extracted["time_logs_24h"] = tables[0]
                if len(tables) >= 2 and not self.base_extracted[
                    "time_logs_morning"
                ]:
                    self.base_extracted["time_logs_morning"] = tables[1]

    def _find_all_time_tables(
        self, sheet_name: str
    ) -> List[List[Dict]]:
        """Find all time log tables in a sheet"""
        cells = self.cell_cache.get(sheet_name, {})
        if not cells:
            return []

        header_rows = []
        checked_rows = set()

        for (r, c), val in sorted(
            cells.items(), key=lambda x: (x[0][0], x[0][1])
        ):
            if r in checked_rows:
                continue
            txt = re.sub(r"[^a-z0-9]", "", str(val).lower())
            if txt not in ("from", "timefrom", "starttime", "timein"):
                continue

            row_texts = {}
            for (r2, c2), v2 in cells.items():
                if r2 == r:
                    row_texts[c2] = str(v2).lower().strip()

            normalized_headers = [re.sub(r"[^a-z0-9]", "", t) for t in row_texts.values()]
            has_to = any(t in ("to", "timeto", "endtime", "timeout") for t in normalized_headers)
            has_hrs = any(
                any(w in t for w in ["hrs", "hour", "duration"])
                for t in normalized_headers
            )

            if has_to and has_hrs:
                header_rows.append((r, row_texts))
                checked_rows.add(r)

        if not header_rows:
            return []

        all_tables = []

        for idx, (header_row, row_texts) in enumerate(header_rows):
            col_map = self._build_time_col_map(row_texts)
            if "from" not in col_map:
                continue

            max_row = (
                header_rows[idx + 1][0] - 1
                if idx + 1 < len(header_rows)
                else header_row + 80
            )
            logs = self._parse_time_table(
                cells, header_row, max_row, col_map
            )
            if logs:
                all_tables.append(logs)

        return all_tables

    def _build_time_col_map(
        self, row_texts: Dict[int, str]
    ) -> Dict[str, int]:
        col_map = {}
        for c2, txt2 in sorted(row_texts.items()):
            normalized = re.sub(r"[^a-z0-9]", "", txt2)
            if normalized in ("from", "timefrom", "starttime", "timein"):
                col_map["from"] = c2
            elif normalized in ("to", "timeto", "endtime", "timeout"):
                col_map["to"] = c2
            elif any(
                w in normalized for w in ["hrs", "duration", "hour"]
            ):
                col_map["hrs"] = c2
            elif any(
                w in txt2 for w in ["main phase", "phase"]
            ):
                col_map["phase"] = c2
            elif (
                normalized in ("code", "maincode", "activitycode", "phasecode", "mainactivitycode", "mainactivity")
                or "maincode" in normalized
                or "activitycode" in normalized
            ):
                col_map["code"] = c2
            elif (
                ("sub" in txt2 or "secondary" in txt2)
                and "code" in txt2
            ):
                col_map["sub_code"] = c2
            elif txt2 == "status":
                col_map["status"] = c2
            elif any(
                w in txt2
                for w in ["npt", "non productive", "trouble"]
            ):
                col_map["npt"] = c2
            elif any(
                w in txt2
                for w in [
                    "rig activity", "description",
                    "activity", "operation",
                ]
            ):
                col_map["desc"] = c2
            elif any(
                w in txt2
                for w in ["contractor", "attributed", "responsible"]
            ):
                col_map["contractor"] = c2
            elif txt2 in ("remark", "remarks", "note", "notes"):
                col_map["remark"] = c2
        return col_map

    def _parse_time_table(
        self,
        cells: Dict,
        header_row: int,
        max_row: int,
        col_map: Dict[str, int],
    ) -> List[Dict]:
        logs = []

        for r in range(header_row + 1, max_row + 1):
            row_cells = {}
            for (r2, c2), v in cells.items():
                if r2 == r:
                    row_cells[c2] = v

            if not row_cells:
                continue

            first_text = str(
                list(row_cells.values())[0]
            ).lower().strip()
            if "total" in first_text or first_text in ("from", "from:"):
                break

            time_from_raw = row_cells.get(col_map.get("from"))
            hrs_raw = row_cells.get(col_map.get("hrs"))

            if not time_from_raw and not hrs_raw:
                # continuation row
                if logs and col_map.get("desc"):
                    desc_val = row_cells.get(col_map["desc"])
                    if desc_val:
                        logs[-1]["activity_description"] += (
                            " " + str(desc_val)
                        )
                continue

            hrs = ValueNormalizer.to_float(hrs_raw) or 0.0

            # Main/sub code.  Some DDR exports leave the dedicated code
            # cells empty and put 2.3 in the phase or description column.
            raw_code = row_cells.get(col_map.get("code", 0))
            raw_sub = row_cells.get(col_map.get("sub_code", 0))
            raw_phase = row_cells.get(col_map.get("phase", 0))
            raw_desc_for_code = row_cells.get(col_map.get("desc", 0))
            code_source = " ".join(str(v or "") for v in (raw_code, raw_sub, raw_phase, raw_desc_for_code))
            composite_match = re.search(r"(?<!\d)(\d{1,2})\s*[./-]\s*(\d{1,2})(?!\d)", code_source)
            if composite_match:
                composite = f"{composite_match.group(1)}.{composite_match.group(2)}"
                if not raw_code or not str(raw_code).strip():
                    raw_code = composite
                if not raw_sub or not str(raw_sub).strip():
                    raw_sub = composite

            if not raw_code or not str(raw_code).strip():
                # Phase names such as "DRL - Drilling" are a valid main-code
                # fallback when the workbook has no Main Code column.
                raw_code = raw_phase
            main_code = CodeResolver.resolve_main_code(raw_code)
            sub_code = CodeResolver.resolve_sub_code(raw_sub, raw_code)
            main_phase = str(raw_phase).strip() if raw_phase else ""

            # NPT
            is_npt = False
            npt_category = ""
            npt_col = col_map.get("npt")
            if npt_col:
                npt_raw = row_cells.get(npt_col)
                if npt_raw is not None:
                    npt_val = str(npt_raw).strip()
                    if (npt_val and npt_val.lower() not in
                        ValueNormalizer.EMPTY_MARKERS | {"0", "no"}):
                        is_npt = True
                        npt_category = npt_val
            else:
                if raw_code:
                    code_str = str(raw_code).lower().strip()
                    npt_prefixes = [
                        "t-", "f-", "w-", "rr-", "rr",
                    ]
                    for prefix in npt_prefixes:
                        if code_str.startswith(prefix):
                            is_npt = True
                            npt_category = str(raw_code).strip()
                            break

            # Description + remark
            raw_desc = row_cells.get(col_map.get("desc", 0))
            description = str(raw_desc).strip() if raw_desc else ""
            if (col_map.get("remark")
                and col_map["remark"] != col_map.get("desc", 0)):
                raw_remark = row_cells.get(
                    col_map.get("remark", 0)
                )
                if raw_remark and str(raw_remark).strip():
                    remark_text = str(raw_remark).strip()
                    if remark_text not in description:
                        description = (
                            f"{description} | {remark_text}"
                            if description
                            else remark_text
                        )

            # Status
            raw_status = row_cells.get(col_map.get("status", 0))
            status = (
                str(raw_status).strip() if raw_status else "PLN"
            )

            # Contractor
            raw_contractor = row_cells.get(
                col_map.get("contractor", 0)
            )
            contractor = (
                str(raw_contractor).strip()
                if raw_contractor else ""
            )
            if is_npt and not contractor:
                contractor = CodeResolver.guess_contractor(
                    npt_category if npt_category else str(
                        raw_code or ""
                    )
                )

            log_entry = {
                "time_from": ValueNormalizer.to_time(time_from_raw),
                "time_to": ValueNormalizer.to_time(
                    row_cells.get(col_map.get("to"))
                ),
                "duration": hrs,
                "main_phase": main_phase,
                "main_code": main_code,
                "sub_code": sub_code,
                "status": status,
                "is_npt": is_npt,
                "npt_category": npt_category,
                "activity_description": description,
                "contractor": contractor,
            }
            logs.append(log_entry)

        return logs

    # ================================================================
    # UI Review & Assign
    # ================================================================
    def _reset_review_table(self):
        self.review_table.setRowCount(0)
        all_fields = []
        for fields in FIELD_GROUPS.values():
            all_fields.extend(fields)

        for fp in all_fields:
            row = self.review_table.rowCount()
            self.review_table.insertRow(row)

            section, key = fp.split(".", 1)
            value = self.final_data.get(section, {}).get(key) \
                if self.final_data else None
            assign = self.assignments.get(fp, {})
            conf = assign.get("confidence", 0) if assign else 0

            if value is not None and conf >= 0.8:
                icon, bg = "🟢", QColor("#eafaf1")
            elif value is not None and conf > 0:
                icon, bg = "🟡", QColor("#fef9e7")
            else:
                icon, bg = "🔴", QColor("#fdedec")

            si = QTableWidgetItem(icon)
            si.setTextAlignment(Qt.AlignCenter)
            si.setBackground(bg)
            self.review_table.setItem(row, 0, si)

            label = FIELD_LABELS.get(fp, fp)
            fi = QTableWidgetItem(label)
            fi.setData(Qt.UserRole, fp)
            fi.setBackground(bg)
            self.review_table.setItem(row, 1, fi)

            val_text = str(value)[:80] if value is not None else "—"
            vi = QTableWidgetItem(val_text)
            vi.setBackground(bg)
            if value is None:
                vi.setForeground(QColor("#ccc"))
            self.review_table.setItem(row, 2, vi)

            ci = QTableWidgetItem(
                f"{conf:.0%}" if conf > 0 else "—"
            )
            ci.setTextAlignment(Qt.AlignCenter)
            ci.setBackground(bg)
            ci.setForeground(
                QColor("#27ae60") if conf >= 0.8
                else QColor("#f39c12") if conf >= 0.6
                else QColor("#e74c3c")
            )
            self.review_table.setItem(row, 3, ci)

        self._update_summary()

    def _update_summary(self):
        total = sum(len(f) for f in FIELD_GROUPS.values())
        detected = len([
            a for a in self.assignments.values()
            if a.get("value") not in [None, ""]
        ])
        high = sum(
            1 for a in self.assignments.values()
            if a.get("confidence", 0) >= 0.8
        )
        tl_24 = len(self.base_extracted.get("time_logs_24h", []))
        tl_m = len(self.base_extracted.get("time_logs_morning", []))
        missing = total - detected

        self.summary_label.setText(
            f"📊 {detected}/{total} detected | "
            f"🟢 {high} high | "
            f"🔴 {missing} missing | "
            f"📋 Logs: {tl_24}+{tl_m}"
        )
        logs = self.base_extracted.get("time_logs_24h", []) + self.base_extracted.get("time_logs_morning", [])
        with_main = sum(bool(log.get("main_code")) for log in logs)
        with_sub = sum(bool(log.get("sub_code")) for log in logs)
        examples = []
        for log in logs:
            if log.get("main_code") or log.get("sub_code"):
                examples.append(f"{log.get('main_code', '')} / {log.get('sub_code', '')}")
            if len(examples) == 3:
                break
        detail = "; ".join(examples)
        self.code_detection_label.setText(
            f"🏷️ Activity codes: Main {with_main}/{len(logs)}, "
            f"Sub {with_sub}/{len(logs)}" + (f" — {detail}" if detail else " — no code columns detected")
        )
        self.import_btn.setEnabled(detected > 0)

    def _filter_review(self):
        f = self.filter_combo.currentText()
        for row in range(self.review_table.rowCount()):
            icon = self.review_table.item(row, 0).text()
            fp = (
                self.review_table.item(row, 1).data(Qt.UserRole)
                or ""
            )
            show = True
            if f == "🟢 High Conf.":
                show = icon == "🟢"
            elif f == "🟡 Medium":
                show = icon == "🟡"
            elif f == "🔴 Missing":
                show = icon == "🔴"
            elif f == "Well":
                show = fp.startswith("well_info.")
            elif f == "Report":
                show = fp.startswith("daily_report.")
            elif f == "Mud":
                show = fp.startswith("mud_report.")
            elif f == "Drilling":
                show = fp.startswith("drilling_params.")
            self.review_table.setRowHidden(row, not show)

    def _on_review_clicked(self, item):
        fi = self.review_table.item(item.row(), 1)
        if fi:
            self._selected_field = fi.data(Qt.UserRole)
            val = self.review_table.item(item.row(), 2).text()
            self.cell_info.setText(
                f"Selected: {fi.text()} = {val}\n"
                f"Click a cell in Excel preview to assign."
            )
            self.assign_btn.setEnabled(True)

    def _on_excel_clicked(self, row: int, col: int):
        sheet = self.current_sheet
        r = row + 1
        c = col + 1
        val = self.cell_cache.get(sheet, {}).get((r, c))
        col_l = get_column_letter(c)

        if self._range_mode:
            self._range_cells.append({
                "sheet": sheet, "row": r, "col": c,
                "val": val, "ref": f"{col_l}{r}",
            })
            refs = " + ".join(x["ref"] for x in self._range_cells)
            vals = " | ".join(
                str(x["val"])[:20] for x in self._range_cells
            )
            self.cell_info.setText(
                f"📐 Range: {refs}\nValues: {vals}"
            )
            self._pending_cell = (sheet, r, c, val)
        else:
            self._pending_cell = (sheet, r, c, val)
            ft = ""
            if self._selected_field:
                ft = f"  →  {FIELD_LABELS.get(self._selected_field, '')}"
            vd = str(val)[:100] if val is not None else "(empty)"
            self.cell_info.setText(
                f"📍 {sheet}!{col_l}{r}: {vd}{ft}"
            )
            self.assign_btn.setEnabled(True)

    def _assign_cell(self):
        if not self._selected_field:
            QMessageBox.warning(
                self, "No Field",
                "Select a field first (click on review table).",
            )
            return
        if not self._pending_cell:
            QMessageBox.warning(
                self, "No Cell",
                "Click a cell in the Excel preview first.",
            )
            return

        fp = self._selected_field

        if self._range_mode and self._range_cells:
            val = self._build_range_value()
            ref_text = " + ".join(
                c["ref"] for c in self._range_cells
            )
            self.assignments[fp] = {
                "sheet": self._range_cells[0]["sheet"],
                "row": self._range_cells[0]["row"],
                "col": self._range_cells[0]["col"],
                "value": str(val)[:100] if val else "",
                "confidence": 1.0,
                "manual": True,
                "range": [
                    {"row": rc["row"], "col": rc["col"]}
                    for rc in self._range_cells
                ],
            }
            self._range_cells = []
        else:
            sheet, r, c, val = self._pending_cell
            col_l = get_column_letter(c)
            ref_text = f"{sheet}!{col_l}{r}"

            # record for learning
            self._record_learning(fp, sheet, r, c)

            self.assignments[fp] = {
                "sheet": sheet,
                "row": r,
                "col": c,
                "value": str(val)[:100] if val is not None else "",
                "confidence": 1.0,
                "manual": True,
            }

        self.final_data = self._build_final_data_from_assignments()
        self._reset_review_table()
        self._display_sheet(self.current_sheet)

        label = FIELD_LABELS.get(fp, fp)
        self.cell_info.setText(
            f"✅ {ref_text} → {label}: {str(val)[:40]}"
        )
        self.assign_btn.setEnabled(False)
        self._pending_cell = None

    def _record_learning(
        self, field_path: str, sheet: str, v_row: int, v_col: int
    ):
        """Record a manual correction for future learning"""
        cells = self.cell_cache.get(sheet, {})
        pattern = FIELD_PATTERNS.get(field_path, {})
        all_kw = (
            pattern.get("keywords", [])
            + pattern.get("synonyms", [])
        )

        # find nearest label cell
        best_anchor = None
        best_dist = 999
        for (r, c), val in cells.items():
            if r == v_row and c == v_col:
                continue
            cell_text = str(val).strip().lower()
            for kw in all_kw:
                if kw.lower() in cell_text or cell_text in kw.lower():
                    dist = abs(r - v_row) + abs(c - v_col)
                    if dist < best_dist:
                        best_dist = dist
                        best_anchor = (cell_text, r, c)
                    break

        if best_anchor:
            anchor_text, ar, ac = best_anchor
            self.learning.record_correction(
                field_path,
                anchor_text,
                "relative",
                v_row - ar,
                v_col - ac,
            )

    def _build_range_value(self):
        vals = [x["val"] for x in self._range_cells]
        cleaned = [
            str(v).strip() for v in vals
            if v is not None and str(v).strip()
        ]

        # try date
        if self._selected_field and "date" in self._selected_field.lower():
            result = ValueNormalizer.combine_date_parts(cleaned)
            if result:
                return result

        return " ".join(cleaned)

    def _toggle_range_mode(self, checked):
        self._range_mode = checked
        self._range_cells = []
        if checked:
            self.range_btn.setText("📐 Range Mode: ON")
            self.range_btn.setStyleSheet(
                "background: #e74c3c; color: white; "
                "font-weight: bold; font-size: 9px; "
                "border-radius: 3px; border: none;"
            )
            self.excel_table.setSelectionMode(
                QTableWidget.MultiSelection
            )
        else:
            self.range_btn.setText("📐 Range Mode: OFF")
            self.range_btn.setStyleSheet(
                "background: #7f8c8d; color: white; "
                "font-weight: bold; font-size: 9px; "
                "border-radius: 3px; border: none;"
            )
            self.excel_table.setSelectionMode(
                QTableWidget.SingleSelection
            )

    # ================================================================
    # Build Final Data
    # ================================================================
    def _build_final_data_from_assignments(self) -> Dict:
        result = {
            "well_info": dict(
                self.base_extracted.get("well_info", {})
            ),
            "daily_report": dict(
                self.base_extracted.get("daily_report", {})
            ),
            "mud_report": dict(
                self.base_extracted.get("mud_report", {})
            ),
            "drilling_params": dict(
                self.base_extracted.get("drilling_params", {})
            ),
            "time_logs_24h": list(
                self.base_extracted.get("time_logs_24h", [])
            ),
            "time_logs_morning": list(
                self.base_extracted.get("time_logs_morning", [])
            ),
            "metadata": dict(
                self.base_extracted.get("metadata", {})
            ),
        }

        for fp, assign in self.assignments.items():
            if "." not in fp:
                continue
            section, key = fp.split(".", 1)
            if section not in result:
                result[section] = {}

            raw_value = assign.get("value", "")
            clean_value = self._clean_value_for_field(fp, raw_value)
            if clean_value is not None:
                result[section][key] = clean_value

        return result

    def _clean_value_for_field(self, field_path: str, value):
        if value is None:
            return None

        # check if it's a label
        if self.detector and self.detector._is_label(
            str(value), field_path
        ):
            return None

        pattern = FIELD_PATTERNS.get(field_path, {})
        data_type = pattern.get("type", "str")
        valid_range = pattern.get("valid_range")

        return ValueNormalizer.convert(value, data_type, valid_range)

    # ================================================================
    # Template Save/Load (Anchor-based v2)
    # ================================================================
    def _save_template(self):
        name = self.template_name.text().strip()
        if not name:
            name = f"Template_{datetime.now().strftime('%Y%m%d_%H%M')}"
        if not self.assignments:
            QMessageBox.warning(
                self, "No Data", "Detect or assign fields first."
            )
            return

        template = {
            "name": name,
            "version": "2.0",
            "created": datetime.now().isoformat(),
            "assignments": {},
        }

        for fp, assign in self.assignments.items():
            sheet = assign.get("sheet", "")
            row = assign.get("row", 0)
            col = assign.get("col", 0)

            # find anchor text (label near the value)
            anchor_text = self._find_anchor_for_cell(
                sheet, row, col, fp
            )

            entry = {
                "sheet": sheet,
                "row": row,
                "col": col,
                "anchor_text": anchor_text,
                "value_sample": str(assign.get("value", ""))[:50],
                "type": "single",
            }

            if assign.get("range"):
                entry["type"] = "range"
                entry["range_cells"] = assign["range"]

            template["assignments"][fp] = entry

        # Save
        td = os.path.join(
            os.path.dirname(os.path.dirname(__file__)),
            "templates",
        )
        os.makedirs(td, exist_ok=True)
        safe_name = re.sub(r'[^\w\-]', '_', name)
        filepath = os.path.join(td, f"{safe_name}.json")

        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(
                template, f, indent=2,
                ensure_ascii=False, default=str,
            )

        QMessageBox.information(
            self, "✅ Saved",
            f"Template '{name}' saved!\n"
            f"Fields: {len(self.assignments)}\n"
            f"Format: Anchor-based v2.0",
        )

    def _find_anchor_for_cell(
        self, sheet: str, value_row: int, value_col: int,
        field_path: str,
    ) -> str:
        """Find the best anchor label near a value cell"""
        cells = self.cell_cache.get(sheet, {})
        if not cells:
            return ""

        pattern = FIELD_PATTERNS.get(field_path, {})
        all_kw = (
            pattern.get("keywords", [])
            + pattern.get("synonyms", [])
        )

        best_anchor = ""
        best_score = 0.0

        # search in radius around value
        for dr in range(-5, 3):
            for dc in range(-15, 3):
                r = value_row + dr
                c = value_col + dc
                if r == value_row and c == value_col:
                    continue
                val = cells.get((r, c))
                if val is None:
                    continue

                cell_text = str(val).strip()
                cell_lower = cell_text.lower()

                for kw in all_kw:
                    score = SequenceMatcher(
                        None, cell_lower, kw.lower()
                    ).ratio()
                    if score > best_score and score > 0.5:
                        best_score = score
                        best_anchor = cell_text

        return best_anchor

    def _load_template_file(self):
        td = os.path.join(
            os.path.dirname(os.path.dirname(__file__)),
            "templates",
        )
        if not os.path.exists(td):
            return
        fp, _ = QFileDialog.getOpenFileName(
            self, "Load Template", td, "JSON Files (*.json)",
        )
        if not fp:
            return
        with open(fp, 'r', encoding='utf-8') as f:
            tmpl = json.load(f)
        self._apply_template(tmpl)

    def _apply_template(self, tmpl: dict):
        """
        Apply template - supports both v1 (coordinate) and v2 (anchor)
        """
        assigns = tmpl.get("assignments", {})
        if not assigns:
            return

        version = tmpl.get("version", "1.0")
        self.assignments.clear()
        self.confidence_scores.clear()

        for fp, info in assigns.items():
            value = None
            conf = 0.0

            if version >= "2.0" and info.get("anchor_text"):
                # Anchor-based: find anchor first
                result = self._find_by_anchor(fp, info)
                if result:
                    value, sheet, r, c = result
                    conf = 0.95
                else:
                    # fallback to coordinates
                    result = self._find_by_coordinates(fp, info)
                    if result:
                        value, sheet, r, c = result
                        conf = 0.70
            else:
                # v1: coordinate-based
                result = self._find_by_coordinates(fp, info)
                if result:
                    value, sheet, r, c = result
                    conf = 0.90

            if value is not None:
                self.assignments[fp] = {
                    "sheet": sheet,
                    "row": r,
                    "col": c,
                    "value": str(value)[:100],
                    "confidence": conf,
                    "template": True,
                }
                self.confidence_scores[fp] = conf

        self.final_data = self._build_final_data_from_assignments()
        self._reset_review_table()
        if self.current_sheet:
            self._display_sheet(self.current_sheet)

    def _find_by_anchor(
        self, field_path: str, info: dict
    ) -> Optional[Tuple]:
        """Find value by anchor text (shift-resistant)"""
        anchor = info.get("anchor_text", "").lower().strip()
        if not anchor:
            return None

        orig_sheet = info.get("sheet", "")
        orig_row = info.get("row", 0)
        orig_col = info.get("col", 0)

        # calculate original offset from anchor to value
        # we need to find the anchor in original and compute offset
        # for now use pattern-based search

        pattern = FIELD_PATTERNS.get(field_path, {})
        data_type = pattern.get("type", "str")
        valid_range = pattern.get("valid_range")

        # search in all sheets, prefer original
        sheets_order = []
        if orig_sheet in self.cell_cache:
            sheets_order.append(orig_sheet)
        for sn in self.cell_cache:
            if sn not in sheets_order:
                sheets_order.append(sn)

        for sheet_name in sheets_order:
            cells = self.cell_cache.get(sheet_name, {})
            for (r, c), val in cells.items():
                cell_text = str(val).strip().lower()
                if anchor not in cell_text and cell_text not in anchor:
                    # fuzzy
                    ratio = SequenceMatcher(
                        None, cell_text, anchor
                    ).ratio()
                    if ratio < 0.75:
                        continue

                # found anchor - search for value nearby
                # try right first
                for dc in range(1, 16):
                    v = cells.get((r, c + dc))
                    if v is None:
                        continue
                    if ValueNormalizer.is_empty(v):
                        continue
                    if self.detector and self.detector._is_label(str(v)):
                        continue
                    converted = ValueNormalizer.convert(
                        v, data_type, valid_range
                    )
                    if converted is not None:
                        return (converted, sheet_name, r, c + dc)

                # try below
                for dr in range(1, 4):
                    for dc2 in range(0, 6):
                        v = cells.get((r + dr, c + dc2))
                        if v is None:
                            continue
                        if ValueNormalizer.is_empty(v):
                            continue
                        if self.detector and self.detector._is_label(str(v)):
                            continue
                        converted = ValueNormalizer.convert(
                            v, data_type, valid_range
                        )
                        if converted is not None:
                            return (
                                converted, sheet_name,
                                r + dr, c + dc2,
                            )

        return None

    def _find_by_coordinates(
        self, field_path: str, info: dict
    ) -> Optional[Tuple]:
        """Fallback: find value by exact coordinates"""
        sheet = info.get("sheet", "")
        r = info.get("row", 0)
        c = info.get("col", 0)

        val = self.cell_cache.get(sheet, {}).get((r, c))
        if val is not None:
            return (val, sheet, r, c)

        # try nearby cells (±2)
        cells = self.cell_cache.get(sheet, {})
        for dr in range(-2, 3):
            for dc in range(-2, 3):
                if dr == 0 and dc == 0:
                    continue
                v = cells.get((r + dr, c + dc))
                if v is not None and not ValueNormalizer.is_empty(v):
                    return (v, sheet, r + dr, c + dc)

        return None

    # ================================================================
    # Import
    # ================================================================
    def _do_import(self):
        if not self.assignments:
            return
        self.final_data = self._build_final_data_from_assignments()
        self.import_completed.emit(self.final_data)
        self.accept()

    def get_final_data(self) -> Dict:
        return self.final_data or self._build_final_data_from_assignments()