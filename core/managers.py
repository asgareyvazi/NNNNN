# core/managers.py
"""
Core Managers - مدیرهای مرکزی (نسخه اصلاح شده)
- حذف متدهای اشتباهی که به کلاس اشتباه رفته بودند
- رفع show_warning تکراری
- رفع TableManager._setup_table_optimizations
"""
import logging
from typing import Dict, Any, List
from datetime import datetime
import os
import csv

from PySide6.QtCore import *
from PySide6.QtWidgets import *
from PySide6.QtGui import *

logger = logging.getLogger(__name__)


# ==================== StatusBar Manager ====================
class StatusBarManager(QObject):
    """مدیریت StatusBar - Singleton"""

    show_message_signal = Signal(str, str, int)
    clear_message_signal = Signal(str)
    show_progress_signal = Signal(str, str)
    show_success_signal = Signal(str, str)
    show_error_signal = Signal(str, str)

    _instance = None

    def __new__(cls):
        if cls._instance is None:
            cls._instance = super().__new__(cls)
        return cls._instance

    def __init__(self):
        if not hasattr(self, "_initialized"):
            super().__init__()
            self._widgets = {}
            self._initialized = True

            self.show_message_signal.connect(self._handle_show_message)
            self.clear_message_signal.connect(self._handle_clear_message)
            self.show_progress_signal.connect(self._handle_show_progress)
            self.show_success_signal.connect(self._handle_show_success)
            self.show_error_signal.connect(self._handle_show_error)

    def _handle_show_message(self, widget_name: str, message: str, timeout: int):
        """✅ FIX: timeout واقعاً استفاده شود"""
        try:
            if widget_name in self._widgets:
                widget = self._widgets[widget_name]
                if isinstance(widget, QMainWindow):
                    if widget.statusBar():
                        widget.statusBar().showMessage(message, timeout)
                elif hasattr(widget, 'status_bar') and widget.status_bar:
                    widget.status_bar.showMessage(message, timeout)
                else:
                    logger.debug(f"[{widget_name}]: {message}")
        except Exception as e:
            print(f"StatusBar error for {widget_name}: {e}")

    def _handle_clear_message(self, widget_name: str):
        try:
            if widget_name in self._widgets:
                widget = self._widgets[widget_name]
                if isinstance(widget, QMainWindow):
                    widget.statusBar().clearMessage()
        except Exception as e:
            logger.error(f"Error clearing message for {widget_name}: {e}")

    def _handle_show_progress(self, widget_name: str, message: str):
        self.show_message_signal.emit(widget_name, f"🔄 {message}...", 0)

    def _handle_show_success(self, widget_name: str, message: str):
        self.show_message_signal.emit(widget_name, f"✅ {message}", 3000)

    def _handle_show_error(self, widget_name: str, message: str):
        self.show_message_signal.emit(widget_name, f"❌ {message}", 5000)

    def register_widget(self, widget_name: str, widget: QWidget):
        self._widgets[widget_name] = widget
        # QObject.destroyed is the reliable lifecycle hook; it also prevents
        # stale widgets from being retained by the process-wide manager.
        try:
            widget.destroyed.connect(lambda *_: self.unregister_widget(widget_name))
        except (AttributeError, RuntimeError):
            pass
        logger.debug(f"📝 Registered widget: {widget_name}")

    def unregister_widget(self, widget_name: str):
        self._widgets.pop(widget_name, None)

    def register_main_window(self, main_window: QMainWindow):
        self.register_widget("MainWindow", main_window)

    def show_message(self, widget_name: str, message: str, timeout: int = 3000):
        self.show_message_signal.emit(widget_name, message, timeout)

    def clear_message(self, widget_name: str):
        self.clear_message_signal.emit(widget_name)

    def show_progress(self, widget_name: str, message: str):
        self.show_progress_signal.emit(widget_name, message)

    def show_success(self, widget_name: str, message: str):
        self.show_success_signal.emit(widget_name, message)

    def show_error(self, widget_name: str, message: str):
        self.show_error_signal.emit(widget_name, message)

    def show_warning(self, widget_name: str, message: str):
        self.show_message_signal.emit(widget_name, f"⚠️ {message}", 4000)


# ==================== AutoSave Manager ====================
class AutoSaveManager:
    """مدیریت Auto-Save با dirty flag - نسخه اصلاح شده"""

    # ✅ Constants
    DEFAULT_INTERVAL_MIN = 5
    MIN_INTERVAL_MIN = 1
    MAX_INTERVAL_MIN = 60

    def __init__(self):
        self._timers: Dict[str, QTimer] = {}
        self._widgets: Dict[str, QWidget] = {}
        self._dirty: Dict[str, bool] = {}
        self._enabled = True
        self._status_manager = StatusBarManager()
        self._save_count = 0 

    def enable_for_widget(
        self,
        widget_name: str,
        widget: QWidget,
        interval_minutes: int = DEFAULT_INTERVAL_MIN
    ) -> None:
        """فعال‌سازی auto-save برای یک widget"""
        if not hasattr(widget, "save_data"):
            logger.warning(
                f"Widget '{widget_name}' has no save_data method"
            )
            return

        interval_minutes = max(
            self.MIN_INTERVAL_MIN,
            min(self.MAX_INTERVAL_MIN, interval_minutes)
        )
        interval_ms = interval_minutes * 60 * 1000

        if widget_name in self._timers:
            self._timers[widget_name].stop()

        timer = QTimer()
        timer.timeout.connect(
            lambda: self._auto_save(widget_name, widget)
        )
        timer.start(interval_ms)

        self._timers[widget_name] = timer
        self._widgets[widget_name] = widget
        self._dirty[widget_name] = False

        logger.debug(
            f"AutoSave enabled for '{widget_name}' "
            f"(every {interval_minutes} min)"
        )

    def mark_dirty(self, widget_name: str) -> None:
        """علامت‌گذاری تغییر"""
        if widget_name in self._dirty:
            self._dirty[widget_name] = True

    def mark_clean(self, widget_name: str) -> None:
        """علامت‌گذاری تمیز (بعد از ذخیره)"""
        if widget_name in self._dirty:
            self._dirty[widget_name] = False

    def is_dirty(self, widget_name: str) -> bool:
        """بررسی وجود تغییرات ذخیره نشده"""
        return self._dirty.get(widget_name, False)

    def _auto_save(self, widget_name: str, widget: QWidget) -> None:
        """ذخیره خودکار - فقط اگر تغییر داشته"""
        if not self._enabled:
            return

        if not self._dirty.get(widget_name, False):
            logger.debug(f"AutoSave skip '{widget_name}' (no changes)")
            return

        try:
            if hasattr(widget, 'save_data'):
                result = widget.save_data()
                if result:
                    self._dirty[widget_name] = False
                    self._save_count += 1
                    from datetime import datetime
                    self._status_manager.show_message(
                        widget_name,
                        f"💾 Auto-saved at {datetime.now():%H:%M:%S}",
                        2000
                    )
                    logger.debug(
                        f"AutoSave success for '{widget_name}' "
                        f"(total: {self._save_count})"
                    )
        except Exception as e:
            logger.error(f"AutoSave error for '{widget_name}': {e}")

    def set_enabled(self, enabled: bool) -> None:
        """فعال/غیرفعال کردن همه timer ها"""
        self._enabled = enabled
        for widget_name, timer in self._timers.items():
            if enabled:
                if not timer.isActive():
                    timer.start()
            else:
                timer.stop()

    @property
    def save_count(self) -> int:
        """تعداد کل ذخیره‌های خودکار"""
        return self._save_count

    def __del__(self):
        try:
            for timer in self._timers.values():
                try:
                    if timer and timer.isActive():
                        timer.stop()
                except RuntimeError:
                    pass  # Qt object already deleted - ignore
        except Exception:
            pass

# ==================== Shortcut Manager ====================
class ShortcutManager:
    """مدیریت کلیدهای میانبر"""

    def __init__(self, parent: QWidget):
        self.parent = parent
        self.shortcuts: Dict[str, Dict[str, Any]] = {}
        self._status_manager = StatusBarManager()

    def add_shortcut(self, key_sequence: str, slot, description: str = ""):
        shortcut = QShortcut(QKeySequence(key_sequence), self.parent)
        shortcut.activated.connect(slot)
        self.shortcuts[key_sequence] = {
            "shortcut": shortcut,
            "description": description,
            "slot": slot,
        }
        return shortcut

    def add_shortcut_with_feedback(self, key_sequence: str, slot, description: str = ""):
        def wrapped_slot():
            slot()
            self._status_manager.show_message(
                self.parent.__class__.__name__,
                f"Shortcut: {description}",
                1000
            )
        return self.add_shortcut(key_sequence, wrapped_slot, description)

    def setup_default_shortcuts(self):
        shortcuts = [
            ("Ctrl+S", self._save_current, "Save"),
            ("Ctrl+Shift+S", self._save_all, "Save All"),
            ("F5", self._refresh, "Refresh"),
            ("Ctrl+E", self._export, "Export"),
            ("F1", self._help, "Help"),
        ]
        for key_seq, slot, desc in shortcuts:
            self.add_shortcut_with_feedback(key_seq, slot, desc)

    def _save_current(self):
        if hasattr(self.parent, 'save_current'):
            self.parent.save_current()

    def _save_all(self):
        if hasattr(self.parent, 'save_all'):
            self.parent.save_all()

    def _refresh(self):
        if hasattr(self.parent, 'refresh'):
            self.parent.refresh()

    def _export(self):
        if hasattr(self.parent, 'export'):
            self.parent.export()

    def _help(self):
        if hasattr(self.parent, 'show_help'):
            self.parent.show_help()


# ==================== Table Manager ====================
class TableManager:
    """مدیریت پیشرفته جدول‌ها"""

    def __init__(self, table_widget: QTableWidget, parent_widget=None):
        self.table = table_widget
        self.parent = parent_widget
        self.undo_stack = []
        self.redo_stack = []
        self.max_stack_size = 100
        self._is_recording = True

        self._setup_table_optimizations()
        self.table.itemChanged.connect(self._on_item_changed)
        self._setup_shortcuts()

    def _setup_table_optimizations(self):
        # Stretch mode - فقط Interactive تنظیم میشه
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.table.setWordWrap(True)

        # Scroll
        self.table.setHorizontalScrollMode(QAbstractItemView.ScrollPerPixel)
        self.table.setVerticalScrollMode(QAbstractItemView.ScrollPerPixel)
        self.table.verticalScrollBar().setSingleStep(20)
        self.table.horizontalScrollBar().setSingleStep(20)

        # Appearance
        self.table.setAlternatingRowColors(True)
        self.table.setShowGrid(True)
        self.table.setGridStyle(Qt.SolidLine)
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setSelectionMode(QAbstractItemView.SingleSelection)

        # Row/Column sizes
        self.table.verticalHeader().setDefaultSectionSize(30)
        self.table.verticalHeader().setMinimumSectionSize(25)

    def _setup_shortcuts(self):
        if self.parent:
            QShortcut(QKeySequence("Ctrl+Z"), self.parent).activated.connect(self.undo)
            QShortcut(QKeySequence("Ctrl+Y"), self.parent).activated.connect(self.redo)

    def _on_item_changed(self, item):
        if not self._is_recording or item is None:
            return
        old_value = getattr(item, "_table_manager_old_value", "")
        new_value = item.text()
        if old_value == new_value:
            return
        action = {
            "type": "item_change",
            "row": item.row(),
            "col": item.column(),
            "old_value": old_value,
            "new_value": new_value,
            "timestamp": datetime.now(),
        }
        item._table_manager_old_value = new_value
        self._push_to_undo(action)

    def delete_row(self, row_index=None):
        if row_index is None:
            row_index = self.table.currentRow()
        if row_index < 0 or row_index >= self.table.rowCount():
            return False
        self._is_recording = False
        row_data = []
        for col in range(self.table.columnCount()):
            item = self.table.item(row_index, col)
            row_data.append(item.text() if item else "")
        self.table.removeRow(row_index)
        action = {
            "type": "row_delete",
            "row": row_index,
            "row_data": row_data,
            "timestamp": datetime.now(),
        }
        self._push_to_undo(action)
        self._is_recording = True
        return True

    def add_row(self, default_data=None, position=-1):
        row = self.table.rowCount() if position == -1 else position
        self._is_recording = False
        self.table.insertRow(row)
        if default_data:
            for col, value in enumerate(default_data):
                if col < self.table.columnCount():
                    item = QTableWidgetItem(str(value))
                    item._table_manager_old_value = str(value)
                    item.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)
                    if len(str(value)) > 50:
                        item.setToolTip(str(value))
                    self.table.setItem(row, col, item)
        action = {
            "type": "row_add",
            "row": row,
            "row_data": default_data or [],
            "timestamp": datetime.now(),
        }
        self._push_to_undo(action)
        self._is_recording = True
        return row

    def _push_to_undo(self, action):
        self.undo_stack.append(action)
        if len(self.undo_stack) > self.max_stack_size:
            self.undo_stack.pop(0)
        self.redo_stack.clear()

    def undo(self):
        if not self.undo_stack:
            return False
        self._is_recording = False
        action = self.undo_stack.pop()
        self.redo_stack.append(action)
        if action["type"] == "item_change":
            item = self.table.item(action["row"], action["col"])
            if item:
                item.setText(action["old_value"])
                item._table_manager_old_value = action["old_value"]
        elif action["type"] == "row_add":
            self.table.removeRow(action["row"])
        elif action["type"] == "row_delete":
            self.table.insertRow(action["row"])
            for col, value in enumerate(action["row_data"]):
                if col < self.table.columnCount():
                    item = QTableWidgetItem(str(value))
                    item._table_manager_old_value = str(value)
                    self.table.setItem(action["row"], col, item)
        self._is_recording = True
        return True

    def redo(self):
        if not self.redo_stack:
            return False
        self._is_recording = False
        action = self.redo_stack.pop()
        self.undo_stack.append(action)
        if action["type"] == "item_change":
            item = self.table.item(action["row"], action["col"])
            if item:
                item.setText(action["new_value"])
                item._table_manager_old_value = action["new_value"]
        elif action["type"] == "row_add":
            self.table.insertRow(action["row"])
            for col, value in enumerate(action["row_data"]):
                if col < self.table.columnCount():
                    item = QTableWidgetItem(str(value))
                    item._table_manager_old_value = str(value)
                    self.table.setItem(action["row"], col, item)
        elif action["type"] == "row_delete":
            self.table.removeRow(action["row"])
        self._is_recording = True
        return True

    def can_undo(self):
        return len(self.undo_stack) > 0

    def can_redo(self):
        return len(self.redo_stack) > 0

    def clear_history(self):
        self.undo_stack.clear()
        self.redo_stack.clear()

    def set_alternating_row_colors(self, enabled=True, color1="#FFFFFF", color2="#F5F5F5"):
        self.table.setAlternatingRowColors(enabled)
        if enabled:
            self.table.setStyleSheet(f"""
                QTableWidget {{
                    alternate-background-color: {color2};
                    background-color: {color1};
                    gridline-color: #D0D0D0;
                }}
                QTableWidget::item:selected {{
                    background-color: #0078D4;
                    color: white;
                }}
            """)

    def set_column_stretch_mode(self, mode="stretch"):
        modes = {
            "stretch": QHeaderView.Stretch,
            "interactive": QHeaderView.Interactive,
            "fixed": QHeaderView.Fixed,
            "resize_to_contents": QHeaderView.ResizeToContents,
        }
        self.table.horizontalHeader().setSectionResizeMode(modes.get(mode, QHeaderView.Stretch))

    def import_from_csv(self, filename):
        try:
            with open(filename, 'r', encoding='utf-8-sig') as file:
                reader = csv.reader(file)
                rows = list(reader)
            if not rows:
                return False
            self.table.setRowCount(0)
            col_count = len(rows[0]) if rows else self.table.columnCount()
            self.table.setColumnCount(col_count)
            self.table.setHorizontalHeaderLabels(rows[0])
            for row_idx, row_data in enumerate(rows[1:]):
                self.add_row(row_data)
            return True
        except Exception as e:
            logger.error(f"Error importing from CSV: {e}")
            return False


# ==================== Table Button Manager ====================
class TableButtonManager:
    """مدیریت متمرکز دکمه‌های Add/Remove برای جدول‌ها"""

    @staticmethod
    def add_table_row(table, default_data=None, position=-1):
        if hasattr(table, "_table_manager"):
            return table._table_manager.add_row(default_data, position)
        else:
            row = table.rowCount() if position == -1 else position
            table.insertRow(row)
            if default_data:
                for col, value in enumerate(default_data):
                    if col < table.columnCount():
                        table.setItem(row, col, QTableWidgetItem(str(value)))
            return row

    @staticmethod
    def remove_table_row(table, row_index=None):
        if hasattr(table, "_table_manager"):
            return table._table_manager.delete_row(row_index)
        else:
            if row_index is None:
                row_index = table.currentRow()
            if 0 <= row_index < table.rowCount():
                table.removeRow(row_index)
                return True
            return False


# ==================== Export Manager ====================
class ExportManager:
    """مدیریت export به فرمت‌های مختلف"""

    def __init__(self, parent=None):
        self.parent = parent
        self._status_manager = StatusBarManager()

    def export_table_with_dialog(self, table_widget, default_name="export"):
        if not self.parent:
            return None
        formats = ["CSV", "PDF", "Excel"]
        format_choice, ok = QInputDialog.getItem(
            self.parent, "Export Format", "Select export format:", formats, 0, False
        )
        if not ok or not format_choice:
            return None

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        ext_map = {"CSV": ".csv", "PDF": ".pdf", "Excel": ".xlsx"}
        ext = ext_map.get(format_choice, ".csv")
        default_filename = f"{default_name}_{timestamp}{ext}"
        filter_map = {
            "CSV": "CSV Files (*.csv);;All Files (*.*)",
            "PDF": "PDF Files (*.pdf);;All Files (*.*)",
            "Excel": "Excel Files (*.xlsx);;All Files (*.*)",
        }

        filename, _ = QFileDialog.getSaveFileName(
            self.parent, "Save Export File", default_filename, filter_map.get(format_choice, "")
        )
        if not filename:
            return None
        if not filename.endswith(ext):
            filename += ext

        self._status_manager.show_progress(
            self.parent.__class__.__name__, f"Exporting to {format_choice}..."
        )

        result = self.export_table(table_widget, format_choice.lower(), filename)
        if result:
            self._status_manager.show_success(
                self.parent.__class__.__name__,
                f"Exported to: {os.path.basename(filename)}",
            )
        else:
            self._status_manager.show_error(
                self.parent.__class__.__name__, "Export failed"
            )
        return result

    def export_table(self, table_widget, format="csv", filename=None):
        if format.lower() == "csv":
            return self._export_to_csv(table_widget, filename)
        elif format.lower() == "pdf":
            return self._export_to_pdf(table_widget, filename)
        elif format.lower() == "excel":
            return self._export_to_excel(table_widget, filename)
        return None

    def _export_to_csv(self, table_widget, filename=None):
        try:
            if not filename:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"export_{timestamp}.csv"
            with open(filename, "w", newline="", encoding="utf-8") as file:
                writer = csv.writer(file)
                headers = []
                for col in range(table_widget.columnCount()):
                    header_item = table_widget.horizontalHeaderItem(col)
                    headers.append(header_item.text() if header_item else f"Column {col+1}")
                writer.writerow(headers)
                for row in range(table_widget.rowCount()):
                    row_data = []
                    for col in range(table_widget.columnCount()):
                        item = table_widget.item(row, col)
                        row_data.append(item.text() if item else "")
                    writer.writerow(row_data)
            return filename
        except Exception as e:
            logger.error(f"CSV export failed: {e}")
            return None

    def _export_to_pdf(self, table_widget, filename=None):
        """Export the complete table, not just a title.

        QTextDocument is part of Qt and therefore avoids an optional
        reportlab dependency.  Long tables naturally paginate when printed.
        """
        try:
            from html import escape
            from PySide6.QtGui import QTextDocument
            from PySide6.QtPrintSupport import QPrinter
            if not filename:
                filename = f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            headers = []
            for col in range(table_widget.columnCount()):
                item = table_widget.horizontalHeaderItem(col)
                headers.append(escape(item.text() if item else f"Column {col + 1}"))
            rows = []
            for row in range(table_widget.rowCount()):
                cells = []
                for col in range(table_widget.columnCount()):
                    item = table_widget.item(row, col)
                    cells.append(escape(item.text() if item else ""))
                rows.append("<tr>" + "".join(f"<td>{cell}</td>" for cell in cells) + "</tr>")
            html = """<!doctype html><html><head><meta charset='utf-8'>
            <style>@page{margin:12mm}body{font-family:Arial;font-size:8pt}
            h1{font-size:14pt;color:#2c3e50}table{width:100%;border-collapse:collapse}
            th,td{border:1px solid #bfc5ca;padding:4px;word-wrap:break-word}
            th{background:#2c3e50;color:white}tr:nth-child(even){background:#f4f6f7}
            </style></head><body><h1>Table Export</h1><table><thead><tr>"""
            html += "".join(f"<th>{header}</th>" for header in headers)
            html += "</tr></thead><tbody>" + "".join(rows) + "</tbody></table></body></html>"
            printer = QPrinter(QPrinter.HighResolution)
            printer.setOutputFormat(QPrinter.PdfFormat)
            printer.setOutputFileName(filename)
            doc = QTextDocument()
            doc.setHtml(html)
            doc.print_(printer)
            return filename
        except Exception as e:
            logger.error(f"PDF export failed: {e}", exc_info=True)
            return None

    def _export_to_excel(self, table_widget, filename=None):
        try:
            from openpyxl import Workbook
            if not filename:
                filename = f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "Table Export"
            for col in range(table_widget.columnCount()):
                header = table_widget.horizontalHeaderItem(col)
                ws.cell(row=1, column=col+1,
                        value=header.text() if header else f"Column {col+1}")
            for row in range(table_widget.rowCount()):
                for col in range(table_widget.columnCount()):
                    item = table_widget.item(row, col)
                    ws.cell(row=row+2, column=col+1,
                            value=item.text() if item else "")
            wb.save(filename)
            return filename
        except ImportError:
            logger.error("Excel export requires 'openpyxl' package")
            return None
        except Exception as e:
            logger.error(f"Excel export failed: {e}")
            return None

    def export_rows(self, rows, headers, filename, format="csv", title="Export"):
        """Export database/query rows consistently across all tabs.

        ``rows`` may contain dictionaries or sequences.  This is the common
        path for report, analysis and import-preview exports, so tabs no
        longer need slightly different CSV implementations.
        """
        if not filename or not headers:
            return None
        try:
            normalized = []
            for row in rows or []:
                if isinstance(row, dict):
                    normalized.append([row.get(header, "") for header in headers])
                else:
                    normalized.append(list(row))
            if format.lower() == "csv":
                with open(filename, "w", newline="", encoding="utf-8-sig") as stream:
                    writer = csv.writer(stream)
                    writer.writerow(headers)
                    writer.writerows(normalized)
                return filename
            if format.lower() == "excel":
                from openpyxl import Workbook
                wb = Workbook()
                ws = wb.active
                ws.title = title[:31] or "Export"
                ws.append(list(headers))
                for row in normalized:
                    ws.append(row)
                ws.freeze_panes = "A2"
                ws.auto_filter.ref = ws.dimensions
                wb.save(filename)
                return filename
            if format.lower() == "html":
                from html import escape
                head = "".join(f"<th>{escape(str(h))}</th>" for h in headers)
                body = "".join("<tr>" + "".join(f"<td>{escape(str(v or ''))}</td>" for v in row) + "</tr>" for row in normalized)
                with open(filename, "w", encoding="utf-8") as stream:
                    stream.write(f"<html><head><meta charset='utf-8'><title>{escape(title)}</title></head><body><h1>{escape(title)}</h1><table border='1'><thead><tr>{head}</tr></thead><tbody>{body}</tbody></table></body></html>")
                return filename
        except Exception as e:
            logger.error("Row export failed: %s", e, exc_info=True)
        return None

    def export_image(self, pixmap_or_widget, filename=None):
        if not filename:
            filename = f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png"
        if isinstance(pixmap_or_widget, QPixmap):
            pixmap_or_widget.save(filename)
            return filename
        elif hasattr(pixmap_or_widget, 'grab'):
            pixmap = pixmap_or_widget.grab()
            pixmap.save(filename)
            return filename
        else:
            logger.error("Unsupported type for export_image")
            return None


# ==================== Drilling Manager ====================
class DrillingManager:
    """مدیریت محاسبات حفاری - نسخه اصلاح‌شده"""

    @staticmethod
    def calculate_tfa(nozzles_data: list) -> float:
        import math
        total_tfa = 0.0
        for nozzle in nozzles_data:
            size_32 = nozzle.get('size_32nd', 16)
            quantity = nozzle.get('quantity', 1)
            diameter_inch = size_32 / 32.0
            area = math.pi * (diameter_inch / 2.0) ** 2
            total_tfa += area * quantity
        return round(total_tfa, 4)

    @staticmethod
    def calculate_rop(
        depth_in: float,
        depth_out: float,
        hours_on_bottom: float,
    ) -> float:
        """
        محاسبه ROP واقعی روی بیتم
        
        ⚠️ hours_on_bottom = فقط ساعت روی بیتم
        نه total hours که شامل connections هم می‌شود
        
        مرجع: API RP 13D
        """
        bit_drilled = depth_out - depth_in
        if bit_drilled <= 0 or hours_on_bottom <= 0:
            return 0.0
        return round(bit_drilled / hours_on_bottom, 2)
    
    @staticmethod
    def calculate_hsi(pump_pressure: float, flow_rate: float, bit_size: float) -> float:
        """Hydraulic horsepower per square inch of bit area."""
        if bit_size <= 0 or pump_pressure < 0 or flow_rate < 0:
            return 0.0
        hhp = pump_pressure * flow_rate / 1714.0
        bit_area = 3.141592653589793 * (bit_size ** 2) / 4.0
        return round(hhp / bit_area, 2)

    @staticmethod
    def calculate_annular_velocity(
        flow_rate_gpm: float,
        hole_id_inch: float,
        pipe_od_inch: float = 0.0,
    ) -> float:
        """
        AV (ft/min) = 24.51 × Q(gpm) / (Dh² - Dp²)
        
        اگر pipe_od_inch=0: فرض بر این است که hole_id_inch
        مساحت آنولوس است (سازگاری با کد قبلی)
        """
        if pipe_od_inch > 0:
            area = hole_id_inch ** 2 - pipe_od_inch ** 2
        else:
            area = hole_id_inch ** 2

        if area <= 0:
            return 0.0
        av = (24.51 * flow_rate_gpm) / area
        return round(av, 1)
    
    @staticmethod
    def calculate_bit_revolution(rpm: float, hours: float) -> float:
        """محاسبه تعداد دوران بیت."""
        return round(rpm * hours * 60, 0)

    @staticmethod
    def validate_drilling_data(data: dict) -> dict:
        errors = {}
        required_fields = ['bit_no', 'bit_size', 'depth_in', 'depth_out']
        for field in required_fields:
            if field not in data or not str(data.get(field, '')).strip():
                errors[field] = f"{field.replace('_', ' ').title()} is required"

        numeric_fields = [
            'bit_size', 'depth_in', 'depth_out',
            'wob_min', 'wob_max', 'rpm_min', 'rpm_max',
            'torque_min', 'torque_max'
        ]
        for field in numeric_fields:
            value = data.get(field)
            if value is not None and str(value).strip():
                try:
                    float(value)
                except (ValueError, TypeError):
                    errors[field] = (
                        f"{field.replace('_', ' ').title()} must be a number"
                    )

        depth_in = data.get('depth_in')
        depth_out = data.get('depth_out')
        if depth_in and depth_out:
            try:
                if float(depth_in) >= float(depth_out):
                    errors['depth'] = "Depth Out must be greater than Depth In"
            except (ValueError, TypeError):
                pass

        return errors

# ==================== Quick Setup Functions ====================
def setup_widget_with_managers(widget, widget_name, enable_autosave=True,
                               autosave_interval=5, setup_shortcuts=True):
    """تنظیم سریع ویجت با همه managerها"""
    status_manager = StatusBarManager()
    status_manager.register_widget(widget_name, widget)

    if enable_autosave and hasattr(widget, "save_data"):
        auto_save_manager = AutoSaveManager()
        auto_save_manager.enable_for_widget(widget_name, widget, autosave_interval)

    if setup_shortcuts and hasattr(widget, "setup_shortcuts"):
        widget.setup_shortcuts()

    return True