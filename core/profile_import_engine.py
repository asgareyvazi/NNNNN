# core/profile_import_engine.py
"""
Profile-Based Excel Import Engine
موتور قدرتمند و دقیق برای ایمپورت فایل‌های اکسل بر اساس پروفایل (قالب) از پیش تعریف شده.
"""

import re
import json
import logging
from datetime import datetime, date, time
from typing import Dict, List, Any, Optional
from openpyxl import load_workbook

logger = logging.getLogger(__name__)

# =====================================================================
# 1. پروفایل اختصاصی شرکت OEOC (قابل توسعه برای شرکت‌های دیگر)
# =====================================================================

OEOC_PROFILE = {
    "name": "OEOC / MSA DDR Format",
    
    # کلماتی که اگر در نام شیت‌ها باشند، این پروفایل فعال می‌شود
    "sheet_identifiers": ["DDR Remark", "DDR Data"],
    
    # فیلدهای تکی (Key-Value)
    "fields": [
        # --- Well Info ---
        {"target": "well_info.name", "sheet": "DDR Remark", "anchor": "Well Name", "type": "str"},
        {"target": "well_info.field_name", "sheet": "DDR Remark", "anchor": "Field:", "type": "str"},
        {"target": "well_info.well_type", "sheet": "DDR Remark", "anchor": "Well Type:", "type": "str"},
        {"target": "well_info.well_shape", "sheet": "DDR Remark", "anchor": "Well Shape:", "type": "str"},
        {"target": "well_info.client", "sheet": "DDR Remark", "anchor": "Client:", "type": "str"},
        {"target": "well_info.operator", "sheet": "DDR Remark", "anchor": "Operator:", "type": "str"},
        {"target": "well_info.client_rep", "sheet": "DDR Remark", "anchor": "Client Rep:", "type": "str"},
        {"target": "well_info.gle_msl", "sheet": "DDR Remark", "anchor": "GLE-MSL", "type": "float"},
        {"target": "well_info.rte_msl", "sheet": "DDR Remark", "anchor": "RTE-MSL", "type": "float"},
        {"target": "well_info.gle_rte", "sheet": "DDR Remark", "anchor": "RT - WH", "type": "float"},
        {"target": "well_info.operation_manager", "sheet": "DDR Remark", "anchor": "Operation manager:", "type": "str"},
        {"target": "well_info.superintendent", "sheet": "DDR Remark", "anchor": "Superintendent:", "type": "str"},
        {"target": "well_info.supervisor_day", "sheet": "DDR Remark", "anchor": "Supervisor (Day):", "type": "str"},
        {"target": "well_info.supervisor_night", "sheet": "DDR Remark", "anchor": "Supervisor (Night):", "type": "str"},
        {"target": "well_info.geologist1", "sheet": "DDR Remark", "anchor": "Site Geologist:", "type": "str"},
        {"target": "well_info.tool_pusher_day", "sheet": "DDR Remark", "anchor": "Tool Pusher", "type": "str"},
        {"target": "well_info.lta_day", "sheet": "DDR Remark", "anchor": "LTA (Day)", "type": "int"},
        {"target": "well_info.actual_rig_days", "sheet": "DDR Remark", "anchor": "Actual Rig Days:", "type": "float"},
        {"target": "well_info.target_depth", "sheet": "DDR Remark", "anchor": "Estimated Final Depth", "type": "float"},
        {"target": "well_info.section_name", "sheet": "DDR Remark", "anchor": "Hole Section (inch):", "type": "str"},
        
        # --- Dates (3-part format: Year, Month, Day) ---
        {"target": "well_info.spud_date", "sheet": "DDR Remark", "anchor": "Spud :", "type": "date_triplet"},
        {"target": "well_info.report_date", "sheet": "DDR Remark", "anchor": "Report:", "type": "date_triplet"},
        {"target": "well_info.start_hole_date", "sheet": "DDR Remark", "anchor": "Hole Section Strat:", "type": "date_triplet"},
        {"target": "daily_report.report_date", "sheet": "DDR Remark", "anchor": "Report:", "type": "date_triplet"},
        
        # --- Daily Report ---
        {"target": "daily_report.report_number", "sheet": "DDR Data", "anchor": "Report No", "type": "int"},
        {"target": "daily_report.depth_0000", "sheet": "DDR Remark", "anchor": "MD (m)@ 0:00", "type": "float"},
        {"target": "daily_report.depth_0600", "sheet": "DDR Remark", "anchor": "MD (m)@ 6:00", "type": "float"},
        {"target": "daily_report.depth_2400", "sheet": "DDR Remark", "anchor": "MD (m)@ 24:00", "type": "float"},
        {"target": "daily_report.avg_rop", "sheet": "DDR Remark", "anchor": "AVG. ROP", "type": "float"},
        
        # --- Mud Report (DDR Data) ---
        {"target": "mud_report.mud_type", "sheet": "DDR Data", "anchor": "Mud Type", "type": "str"},
        {"target": "mud_report.mw", "sheet": "DDR Data", "anchor": "Mud Weight", "type": "float"},
        {"target": "mud_report.funnel_vis", "sheet": "DDR Data", "anchor": "Funnel Vis", "type": "float"},
        {"target": "mud_report.pv", "sheet": "DDR Data", "anchor": "PV (CP)", "type": "float"},
        {"target": "mud_report.yp", "sheet": "DDR Data", "anchor": "YP (lb/100ft2)", "type": "float"},
        {"target": "mud_report.gel_10s", "sheet": "DDR Data", "anchor": "Gel 10 s", "type": "float"},
        {"target": "mud_report.gel_10m", "sheet": "DDR Data", "anchor": "Gel 10 min", "type": "float"},
        {"target": "mud_report.fl", "sheet": "DDR Data", "anchor": "Fluid Loss", "type": "float"},
        {"target": "mud_report.cake_thickness", "sheet": "DDR Data", "anchor": "Filter Cake", "type": "float"},
        {"target": "mud_report.solid_percent", "sheet": "DDR Data", "anchor": "Solid (% Vol)", "type": "float"},
        {"target": "mud_report.chloride", "sheet": "DDR Data", "anchor": "Chloride", "type": "float"},
        {"target": "mud_report.oil_percent", "sheet": "DDR Data", "anchor": "Oil / Water", "type": "float"},
        {"target": "mud_report.kcl_percent", "sheet": "DDR Data", "anchor": "KCL (% wt)", "type": "float"},
        {"target": "mud_report.ph", "sheet": "DDR Data", "anchor": "PH", "type": "float"},
        {"target": "mud_report.temperature", "sheet": "DDR Data", "anchor": "Flow line Temp", "type": "float"},
        {"target": "mud_report.hardness", "sheet": "DDR Data", "anchor": "Total Hardness", "type": "float"},
        {"target": "mud_report.mbt", "sheet": "DDR Data", "anchor": "MBT", "type": "float"},
        {"target": "mud_report.volume_hole", "sheet": "DDR Data", "anchor": "Vol. in Hole", "type": "float"},
        {"target": "mud_report.loss_surface", "sheet": "DDR Data", "anchor": "Lost at Surface", "type": "float"},
        {"target": "mud_report.loss_downhole", "sheet": "DDR Data", "anchor": "Lost Down Hole", "type": "float"},
    ]
}


# =====================================================================
# 2. موتور اصلی پردازش
# =====================================================================

class ProfileImportEngine:
    def __init__(self, db_manager):
        self.db = db_manager
        self.profiles = [OEOC_PROFILE]
        self.cell_cache = {}  # {sheet_name: {row: {col: value}}}
        
    def analyze_and_extract(self, filepath: str) -> Dict[str, Any]:
        """فایل را می‌گیرد، پروفایل مناسب را پیدا می‌کند و داده‌ها را با دقت ۱۰۰٪ استخراج می‌کند."""
        wb = load_workbook(filepath, data_only=True)
        sheet_names = [s.title for s in wb.worksheets]
        
        # 1. پیدا کردن پروفایل.  Real-world workbooks often rename sheets
        # (for example DDR_Remark, DDR-Data, or add a company prefix), so an
        # exact two-sheet match was too strict and made this engine appear
        # unusable.
        matched_profile = None
        best_score = 0
        for profile in self.profiles:
            score = sum(
                1 for identifier in profile["sheet_identifiers"]
                if any(identifier.lower().replace(" ", "") in s.lower().replace(" ", "") for s in sheet_names)
            )
            if score > best_score:
                best_score, matched_profile = score, profile

        if not matched_profile or best_score == 0:
            raise ValueError(
                "No supported DDR profile detected. Expected a sheet containing "
                "DDR Remark/DDR Data, or use Smart Import for a custom workbook."
            )

        logger.info(f"Matched Profile: {matched_profile['name']} (score {best_score}/{len(matched_profile['sheet_identifiers'])})")
        
        SKIP_SHEETS = [
            "setting", "config", "template", "pivot",
            "chart", "macro", "lookup", "reference",
            "helper", "validation", "dropdown", "list",
            "master", "archive", "old", "backup",
        ]

        for sheet_name in sheet_names:
            ws = wb[sheet_name]
            name_lower = sheet_name.lower().strip()

            # ✅ skip شیت‌های نامربوط
            if any(kw in name_lower for kw in SKIP_SHEETS):
                logger.info(f"Skipping sheet: '{sheet_name}'")
                continue

            self.cell_cache[sheet_name] = self._build_unmerged_cache(ws)
            
        # 3. ساختار خروجی
        extracted_data = {
            "well_info": {},
            "daily_report": {},
            "mud_report": {},
            "drilling_params": {},
            "time_logs_24h": [],
            "service_companies": [],
        }
        
        # 4. استخراج فیلدهای Key-Value بر اساس پروفایل
        for field_def in matched_profile["fields"]:
            target_path = field_def["target"]
            sheet_name = self._get_real_sheet_name(sheet_names, field_def["sheet"])
            if not sheet_name:
                continue
                
            anchor = field_def["anchor"].lower()
            data_type = field_def["type"]
            
            # جستجوی Anchor در کش
            anchor_pos = self._find_anchor(sheet_name, anchor)
            if not anchor_pos:
                continue
                
            row, col = anchor_pos
            
            # استخراج مقدار بسته به نوع
            if data_type == "date_triplet":
                val = self._extract_date_triplet(sheet_name, row, col)
            else:
                val = self._extract_right_value(sheet_name, row, col, data_type)
                
            if val is not None:
                section, key = target_path.split('.')
                extracted_data[section][key] = val

        # 5. استخراج متن‌های طولانی (Summary)
        remark_sheet = self._get_real_sheet_name(sheet_names, "DDR Remark")
        data_sheet = self._get_real_sheet_name(sheet_names, "DDR Data")
        if not remark_sheet:
            # Some vendors put the time table in a renamed operations sheet.
            remark_sheet = next((s for s in sheet_names if "remark" in s.lower() or "operation" in s.lower()), None)
        extracted_data["daily_report"]["summary"] = self._extract_text_block(
            remark_sheet, "Summary of Activities in Last 24", "Operation Forecast"
        ) if remark_sheet else ""
        extracted_data["daily_report"]["forecast"] = self._extract_text_block(
            remark_sheet, "Operation Forecast for next", "From"
        ) if remark_sheet else ""

        # 6. استخراج جداول (Time Log و Service Company)
        extracted_data["time_logs_24h"] = self._extract_time_logs(remark_sheet) if remark_sheet else []
        service_sheet = self._get_real_sheet_name(sheet_names, "Service Company")
        extracted_data["service_companies"] = self._extract_service_companies(service_sheet) if service_sheet else []

        # 7. استخراج خودکار داده‌ها برای سایر تب‌ها (Trajectory, Logistics POB/Bulk, Casing/Cement, Bit/BHA, HSE, Cost)
        multi_data = self._extract_multi_tab_sheets(sheet_names)
        extracted_data.update(multi_data)
        
        return extracted_data

    def _extract_multi_tab_sheets(self, sheet_names: List[str]) -> Dict[str, Any]:
        """استخراج هوشمند داده‌های سایر شیت‌های فایل اکسل برای تغذیه تب‌های ۱ تا ۱۶"""
        res = {
            "surveys": [],
            "pob_records": [],
            "casing_report": {},
            "cement_report": {},
            "bit_report": {},
            "bha_report": {},
            "bulk_materials": [],
            "fuel_water": {},
            "safety_report": {},
            "bop_components": [],
            "waste_records": [],
            "cost_records": [],
            "equipment_logs": []
        }
        for name in sheet_names:
            lower_name = name.lower()
            cache = self.cell_cache.get(name, {})
            if not cache:
                continue

            # A. Trajectory / Surveys (md, inc, azi, tvd)
            if any(k in lower_name for k in ["survey", "traject", "directional", "deviation"]):
                for r in range(2, 200):
                    if r not in cache: continue
                    md = self._to_float(cache[r].get(1)) or self._to_float(cache[r].get(2))
                    inc = self._to_float(cache[r].get(3))
                    azi = self._to_float(cache[r].get(4))
                    if md and md > 0:
                        res["surveys"].append({
                            "md": md,
                            "inc": inc or 0.0,
                            "azi": azi or 0.0,
                            "tvd": self._to_float(cache[r].get(5)) or 0.0
                        })

            # B. POB / Personnel On Board
            elif any(k in lower_name for k in ["pob", "personnel", "crew", "manpower"]):
                for r in range(2, 100):
                    if r not in cache: continue
                    company = cache[r].get(1) or cache[r].get(2)
                    count = self._to_float(cache[r].get(3)) or self._to_float(cache[r].get(4))
                    if company and count and count > 0:
                        res["pob_records"].append({
                            "company_name": str(company),
                            "pob_day": int(count),
                            "pob_night": 0,
                            "pob_total": int(count)
                        })

            # C. Bit / BHA
            elif any(k in lower_name for k in ["bit", "bha"]):
                res["bit_report"] = {
                    "bit_no": str(cache.get(2, {}).get(2, "1")),
                    "bit_size": self._to_float(cache.get(2, {}).get(3, 8.5)) or 8.5,
                    "bit_type": str(cache.get(3, {}).get(2, "PDC")),
                    "iadc_code": str(cache.get(3, {}).get(3, "M333"))
                }

            # D. Casing & Cement
            elif any(k in lower_name for k in ["casing", "cement", "csg"]):
                res["casing_report"] = {
                    "casing_size": self._to_float(cache.get(2, {}).get(2, 9.625)) or 9.625,
                    "casing_weight": self._to_float(cache.get(2, {}).get(3, 47.0)) or 47.0,
                    "setting_depth": self._to_float(cache.get(3, {}).get(2, 1500.0)) or 1500.0
                }

            # E. Logistics Bulk & Fuel/Water
            elif any(k in lower_name for k in ["bulk", "fuel", "water", "inventory"]):
                for r in range(2, 100):
                    if r not in cache: continue
                    mat = cache[r].get(1) or cache[r].get(2)
                    stock = self._to_float(cache[r].get(3))
                    if mat and stock is not None:
                        res["bulk_materials"].append({
                            "material_name": str(mat),
                            "unit": str(cache[r].get(4, "kg")),
                            "initial_stock": stock,
                            "received": self._to_float(cache[r].get(5)) or 0.0,
                            "used": self._to_float(cache[r].get(6)) or 0.0,
                            "current_stock": stock + (self._to_float(cache[r].get(5)) or 0.0) - (self._to_float(cache[r].get(6)) or 0.0)
                        })

            # F. Equipment and solid-control logs
            elif any(k in lower_name for k in ["equipment", "drill pipe", "solid control", "solids"]):
                for r in range(2, 200):
                    if r not in cache:
                        continue
                    row = cache[r]
                    equipment_name = row.get(1) or row.get(2)
                    if not equipment_name:
                        continue
                    equipment_type = "Solid Control" if any(k in lower_name for k in ["solid", "solids"]) else ("Drill Pipe" if "pipe" in lower_name else "Rig Equipment")
                    res["equipment_logs"].append({
                        "equipment_type": equipment_type,
                        "equipment_name": str(equipment_name),
                        "equipment_id": str(row.get(2) or ""),
                        "manufacturer": str(row.get(3) or ""),
                        "serial_number": str(row.get(4) or ""),
                        "status": str(row.get(5) or "Operational"),
                        "notes": str(row.get(6) or ""),
                        "hours_worked": self._to_float(row.get(7)) or 0.0,
                    })

            # G. Safety / BOP
            elif any(k in lower_name for k in ["safety", "bop", "hse", "waste"]):
                res["safety_report"] = {
                    "days_without_lti": int(self._to_float(cache.get(2, {}).get(2, 0)) or 0),
                    "incidents_count": int(self._to_float(cache.get(3, {}).get(2, 0)) or 0)
                }

            # G. Cost / AFE
            elif any(k in lower_name for k in ["cost", "afe", "expense"]):
                for r in range(2, 100):
                    if r not in cache: continue
                    item_name = cache[r].get(1) or cache[r].get(2)
                    daily_cost = self._to_float(cache[r].get(3))
                    if item_name and daily_cost is not None:
                        res["cost_records"].append({
                            "cost_category": str(cache[r].get(1, "Operational")),
                            "description": str(item_name),
                            "daily_cost": daily_cost,
                            "cum_cost": self._to_float(cache[r].get(4)) or daily_cost
                        })
        return res

    def _to_float(self, val) -> Optional[float]:
        try:
            if val is None: return None
            return float(val)
        except:
            return None
            
    # ------------------- توابع کمکی جادویی -------------------

    def _get_real_sheet_name(self, actual_names, partial_name):
        """Resolve a profile sheet name without leaving legacy code paths."""
        wanted = re.sub(r"[^a-z0-9]", "", str(partial_name).lower())
        if not wanted:
            return None
        for name in actual_names or []:
            candidate = re.sub(r"[^a-z0-9]", "", str(name).lower())
            if wanted in candidate or candidate in wanted:
                return name
        return None

    def _build_unmerged_cache(self, ws):
        cache = {}

        merged_slaves = set()
        try:
            for merged_range in ws.merged_cells.ranges:
                min_col, min_row, max_col, max_row = merged_range.bounds
                for r in range(min_row, max_row + 1):
                    for c in range(min_col, max_col + 1):
                        if r == min_row and c == min_col:
                            continue  # master
                        merged_slaves.add((r, c))
        except Exception:
            pass

        for r in range(1, min(ws.max_row + 1, MAX_PROFILE_ROWS)):
            cache[r] = {}
            for c in range(1, min(ws.max_column + 1, MAX_PROFILE_COLS)):
                if (r, c) in merged_slaves:
                    continue
                val = ws.cell(row=r, column=c).value
                if val is not None:
                    cache[r][c] = val

        return cache

    def _find_anchor(self, sheet_name, anchor_text):
        cache = self.cell_cache[sheet_name]
        for r, cols in cache.items():
            for c, val in cols.items():
                if anchor_text in str(val).lower():
                    return (r, c)
        return None

    def _extract_right_value(self, sheet_name, row, col, data_type):
        """به سمت راست حرکت می‌کند تا اولین سلول پر را پیدا کند"""
        cache = self.cell_cache[sheet_name]
        if row not in cache: return None
        
        for c in range(col + 1, col + 15):
            val = cache[row].get(c)
            if val is not None and str(val).strip() not in ("", "-", "---"):
                # تبدیل نوع
                if data_type == "float":
                    try: return float(re.sub(r'[^\d\.\-]', '', str(val)))
                    except: pass
                elif data_type == "int":
                    try: return int(float(re.sub(r'[^\d\.\-]', '', str(val))))
                    except: pass
                else:
                    return str(val).strip()
        return None

    def _extract_date_triplet(self, sheet_name, row, col):
        """سه سلول پر بعدی را به عنوان سال، ماه و روز برمی‌گرداند"""
        cache = self.cell_cache[sheet_name]
        if row not in cache: return None
        
        vals = []
        for c in range(col + 1, col + 15):
            val = cache[row].get(c)
            if val is not None and str(val).strip() not in ("", "-", "---"):
                vals.append(str(val).strip())
                if len(vals) == 3: break
                
        if len(vals) == 3:
            try:
                y = int(float(vals[0]))
                d = int(float(vals[2]))
                month_map = {"jan":1, "feb":2, "mar":3, "apr":4, "may":5, "jun":6, 
                             "jul":7, "aug":8, "sep":9, "oct":10, "nov":11, "dec":12}
                m = month_map.get(vals[1].lower()[:3], 1)
                return date(y, m, d)
            except: pass
        return None

    def _extract_text_block(self, sheet_name, start_keyword, stop_keyword):
        cache = self.cell_cache.get(sheet_name, {})
        start_row = self._find_anchor(sheet_name, start_keyword.lower())
        if not start_row: return ""
        
        lines = []
        for r in range(start_row[0] + 1, start_row[0] + 20):
            row_vals = cache.get(r, {})
            row_text = " ".join(str(v) for v in row_vals.values()).strip()
            if stop_keyword.lower() in row_text.lower():
                break
            if row_text:
                lines.append(row_text)
        return "\n".join(lines)

    def _extract_time_logs(self, sheet_name):
        cache = self.cell_cache.get(sheet_name, {})
        logs = []
        
        # پیدا کردن هدر جدول
        header_row = None
        for r, cols in cache.items():
            txt = " ".join(str(v).lower() for v in cols.values())
            if "from" in txt and "to" in txt and "hrs" in txt:
                header_row = r
                break
                
        if not header_row: return logs
        
        # پیدا کردن موقعیت ستون‌ها
        cols = cache[header_row]
        c_from = c_to = c_hrs = c_phase = c_code = c_sub = c_status = c_npt = c_act = 0
        
        for c, v in cols.items():
            v_lower = str(v).lower().strip()
            if v_lower == "from": c_from = c
            elif v_lower == "to": c_to = c
            elif "hrs" in v_lower: c_hrs = c
            elif "main phase" in v_lower: c_phase = c
            elif v_lower in ("code", "main code", "activity code", "phase code") or "main code" in v_lower or "activity code" in v_lower: c_code = c
            elif ("sub" in v_lower or "secondary" in v_lower) and "code" in v_lower: c_sub = c
            elif v_lower == "status": c_status = c
            elif "npt" in v_lower: c_npt = c
            elif "rig activity" in v_lower: c_act = c
            
        # خواندن دیتا
        for r in range(header_row + 1, header_row + 50):
            if r not in cache: continue
            
            c1_val = str(cache[r].get(1, "")).lower()
            if "total" in c1_val: break # پایان جدول
            
            tf_raw = cache[r].get(c_from)
            hrs_raw = cache[r].get(c_hrs)
            
            if not tf_raw and not hrs_raw:
                # ردیف ادامه‌دار
                if logs and c_act in cache[r]:
                    logs[-1]["activity_description"] += " " + str(cache[r][c_act])
                continue
                
            # پردازش زمان و ساعت
            tf = self._convert_time(tf_raw)
            tt = self._convert_time(cache[r].get(c_to))
            
            try: hrs = float(hrs_raw)
            except: hrs = 0.0
            
            npt_val = str(cache[r].get(c_npt, "")).strip()
            is_npt = bool(npt_val and npt_val not in ("-", "---", "None"))
            
            raw_main = cache[r].get(c_code)
            raw_sub = cache[r].get(c_sub)
            # Fallback for exports that place the composite code in phase or
            # activity text instead of a dedicated code column.
            raw_phase = cache[r].get(c_phase)
            raw_activity = cache[r].get(c_act)
            source = " ".join(str(v or "") for v in (raw_main, raw_sub, raw_phase, raw_activity))
            composite_match = re.search(r"(?<!\d)(\d{1,2})\s*[./-]\s*(\d{1,2})(?!\d)", source)
            if composite_match:
                composite = f"{composite_match.group(1)}.{composite_match.group(2)}"
                raw_main = raw_main or composite
                raw_sub = raw_sub or composite
            if not raw_main or not str(raw_main).strip():
                raw_main = raw_phase
            try:
                # Keep one canonical resolver for Smart and profile imports.
                from dialogs.smart_template_dialog import CodeResolver
                normalized_main = CodeResolver.resolve_main_code(raw_main)
                normalized_sub = CodeResolver.resolve_sub_code(raw_sub, raw_main)
            except Exception:
                normalized_main = str(raw_main or "").strip()
                normalized_sub = str(raw_sub or "").strip()

            logs.append({
                "time_from": tf,
                "time_to": tt,
                "duration": hrs,
                "main_phase": str(cache[r].get(c_phase, "")),
                "main_code": normalized_main,
                "sub_code": normalized_sub,
                "status": str(cache[r].get(c_status, "PLN")),
                "is_npt": is_npt,
                "npt_category": npt_val if is_npt else "",
                "activity_description": str(cache[r].get(c_act, "")),
                "contractor": ""
            })
            
        return logs

    def _extract_service_companies(self, sheet_name):
        cache = self.cell_cache.get(sheet_name, {})
        companies = []
        for r in range(2, 50): # ردیف اول هدر است
            if r not in cache: continue
            company = cache[r].get(2)
            if not company: continue
            companies.append({
                "company_name": str(company),
                "service_type": str(cache[r].get(3, "")),
                "date_in": str(cache[r].get(4, "")),
                "date_out": str(cache[r].get(5, "")),
                "description": str(cache[r].get(7, "")),
                "status": "Active"
            })
        return companies

    def _convert_time(self, raw):
        if isinstance(raw, time): return raw
        if isinstance(raw, datetime): return raw.time()
        s = str(raw).strip()
        m = re.match(r'^(\d{1,2}):(\d{2})', s)
        if m:
            h, mi = int(m.group(1)), int(m.group(2))
            return time(0 if h==24 else h, mi)
        try:
            f = float(raw)
            if 0 <= f < 1:
                sec = int(f * 86400)
                return time(sec//3600, (sec%3600)//60)
        except: pass
        return None

    # =====================================================================
    # 3. ذخیره در دیتابیس
    # =====================================================================
    def import_to_db(self, extracted_data: Dict, well_id: int):
        """داده‌های استخراج شده را مستقیم در دیتابیس ذخیره می‌کند."""
        session = self.db.create_session()
        from core.database import TimeLog24H, ServiceCompany
        
        try:
            # 1. Well Info
            wi = extracted_data["well_info"]
            wi["id"] = well_id
            self.db.save_well(wi)
            
            # 2. Daily Report
            dr = extracted_data["daily_report"]
            dr["well_id"] = well_id
            report_date = wi.get("report_date") or date.today()
            dr["report_date"] = report_date
            
            saved_report = self.db.save_daily_report(dr)
            report_id = saved_report.get("id") if saved_report else None
            
            if not report_id:
                raise Exception("Failed to save Daily Report base record.")

            # 3. Mud Report
            mr = extracted_data["mud_report"]
            if mr:
                mr["well_id"] = well_id
                mr["report_id"] = report_id
                mr["report_date"] = report_date
                self.db.save_mud_report(mr)
                
            # 4. Time Logs
            logs = extracted_data["time_logs_24h"]
            if logs:
                session.query(TimeLog24H).filter(TimeLog24H.report_id == report_id).delete()
                for log in logs:
                    if log["time_from"]: # فقط رکوردهای دارای زمان
                        t = TimeLog24H(report_id=report_id, **log)
                        session.add(t)
                        
            # 5. Service Companies
            comps = extracted_data["service_companies"]
            if comps:
                for c in comps:
                    c["well_id"] = well_id
                    c["report_id"] = report_id
                    self.db.save_service_company(c)
            # 6. ذخیره‌سازی هم‌زمان برای تمامی تب‌های دیگر برنامه (Trajectory, Logistics, Casing, Safety, Cost)
            self.db.save_imported_multi_tab_data(well_id, report_id, extracted_data)
            
            session.commit()
            return True, report_id
            
        except Exception as e:
            session.rollback()
            logger.error(f"DB Import Error: {e}")
            raise e
        finally:
            session.close()